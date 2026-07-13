import streamlit as st
import pandas as pd
import openpyxl
import io
import re
import datetime
import hashlib
import json
import bcrypt
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

# ── CSV parsing ───────────────────────────────────────────────────────────────

MONTH_ABBR = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}
DAILY_RE = re.compile(r"^\d{1,2}[/-]\d{1,2}[/-]\d{4}")


def is_formula(value) -> bool:
    return isinstance(value, str) and value.strip().startswith("=")


def is_datelike(value) -> bool:
    return isinstance(value, (datetime.datetime, datetime.date))


def parse_csv(file_bytes: bytes) -> pd.DataFrame:
    raw = pd.read_csv(io.BytesIO(file_bytes), header=None, dtype=str, encoding="utf-8-sig")
    df = raw.iloc[2:].reset_index(drop=True)
    df.columns = range(df.shape[1])
    return df


def classify_row(date_str: str):
    """Return ('daily', date) | ('monthly', (year, month)) | (None, None)"""
    if not isinstance(date_str, str):
        return None, None
    date_str = date_str.strip()
    if DAILY_RE.match(date_str):
        raw = date_str[:10].replace("/", "-")
        try:
            d = datetime.datetime.strptime(raw, "%m-%d-%Y").date()
            return "daily", d
        except ValueError:
            return None, None
    parts = date_str.split()
    if len(parts) == 2 and parts[0][:3].lower() in MONTH_ABBR:
        try:
            month = MONTH_ABBR[parts[0][:3].lower()]
            year = int(parts[1])
            return "monthly", (year, month)
        except ValueError:
            pass
    return None, None


def safe_float(val):
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


# Margaritaville's PMS exports an "Occupancy Statistics" .xlsx instead of the
# standard "Business on the Books" CSV every other hotel uses. Confirmed the
# SAME export (same "DATE PRINTED" timestamp, same values) is used for both
# SR and Forecast, matching how one CSV already feeds ROB/SR/Forecast
# together for every other hotel — so this is ONE parser covering every
# field either flow needs, not a separate one per workbook type. It
# normalizes the export into a DataFrame with the exact same column
# positions as parse_csv() (0=date, 1=Rms Sold, 4=OOO, 5=Room Revenue,
# 6=ADR, 7=Grp PU TY, 8=Grp N/PU TY, 9=Grp Rev TY, 15=Trans count,
# 16=Trans Rev) — mapping confirmed against real exports — so
# STRATEGY_CSV_COLS / build_strategy_change_plan / build_forecast_change_plan
# need no changes at all.
MARGARITAVILLE_SOURCE_FIELDS = {
    "rms sold":     1,   # -> Forecast Rooms Sold (both future & actual)
    "ooo rms":      4,   # -> SR ooo_rms
    "room revenue": 5,   # -> Forecast Revenue (actual/past dates)
    "adr":          6,   # -> Forecast ADR OTB (future dates)
    "grp pkup rms": 7,   # -> SR grp_pu_ty
    "grp rem":      8,   # -> SR grp_npu_ty ("remaining" = not yet picked up)
    "grp rm rev":   9,   # -> SR grp_rev_ty
    "trans rms":    15,  # -> SR otb_trans
    "trans rm rev": 16,  # -> SR trans_rev_ty
}


def parse_margaritaville_source(file_bytes: bytes) -> pd.DataFrame:
    """Parse Margaritaville's 'Occupancy Statistics' PMS export (feeds ROB*/
    SR/Forecast — see MARGARITAVILLE_SOURCE_FIELDS). Detects the header row
    and field columns by their text labels — never by color; the source
    file's color-coding was only for human reference while this mapping was
    being worked out, not something to parse at runtime (this app never uses
    cell color to find targets). Skips 'History Total' / 'Forecasted Total' /
    'Total' summary rows and the trailing filter/timestamp/hotel-name rows at
    the bottom of the sheet (any row whose date column doesn't parse as a
    real date).
    * ROB mapping not wired up yet — pending a small tweak to be confirmed.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.worksheets[0]

    header_row = None
    for r in range(1, min(ws.max_row, 30) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "history/forecasted" in v.strip().lower():
                header_row = r
                break
        if header_row:
            break
    if header_row is None:
        raise ValueError("Could not find the 'History/Forecasted' header row in the source file.")

    col_for_field = {}
    for c in range(1, ws.max_column + 1):
        label = str(ws.cell(header_row, c).value or "").strip().lower()
        for field_label, dest_col in MARGARITAVILLE_SOURCE_FIELDS.items():
            if label == field_label:
                col_for_field[dest_col] = c
    missing = [label for label, dest_col in MARGARITAVILLE_SOURCE_FIELDS.items() if dest_col not in col_for_field]
    if missing:
        raise ValueError(f"Could not find expected column(s) in source file: {', '.join(missing)}.")

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        label = str(ws.cell(r, 1).value or "").strip()
        if not label or "total" in label.lower():
            continue
        date_val = ws.cell(r, 2).value
        if not isinstance(date_val, str) or not DAILY_RE.match(date_val.strip()):
            continue
        row_data = {0: date_val.strip()}
        for dest_col, src_col in col_for_field.items():
            row_data[dest_col] = safe_float(ws.cell(r, src_col).value)
        rows.append(row_data)

    if not rows:
        raise ValueError("No daily rows found in source file.")

    max_col = max(max(r.keys()) for r in rows)
    df = pd.DataFrame(rows).reindex(columns=range(max_col + 1))
    return _add_margaritaville_monthly_totals(df)


# Columns build_rob_change_plan reads for a "monthly" row: Revenue, Room
# Nights, Grp PU, Grp N/PU, Grp Rev. Same column positions the standard
# Business on the Books CSV already provides monthly totals for directly —
# Margaritaville's source has no such totals, so they're synthesized here by
# summing the daily rows for each calendar month present in the data.
ROB_MONTHLY_SUM_COLS = [1, 5, 7, 8, 9]


def _add_margaritaville_monthly_totals(df: pd.DataFrame) -> pd.DataFrame:
    """Append one synthetic 'monthly' row (e.g. 'Jul 2026') per calendar
    month present in the daily rows, summing ROB_MONTHLY_SUM_COLS — so
    build_rob_change_plan (which only reads rows classify_row calls
    'monthly') works unchanged, the same way it already does for every other
    hotel's CSV, which provides these totals directly."""
    sums = {}  # (year, month) -> {col: running sum}
    for _, row in df.iterrows():
        date_str = str(row[0]).strip() if row[0] else ""
        kind, d = classify_row(date_str)
        if kind != "daily":
            continue
        key = (d.year, d.month)
        bucket = sums.setdefault(key, {c: 0.0 for c in ROB_MONTHLY_SUM_COLS})
        for c in ROB_MONTHLY_SUM_COLS:
            v = row.get(c)
            if v is not None and not pd.isna(v):
                bucket[c] += v

    if not sums:
        return df

    monthly_rows = []
    for (year, month), bucket in sums.items():
        month_name = datetime.date(year, month, 1).strftime("%b")
        row_data = {0: f"{month_name} {year}"}
        row_data.update(bucket)
        monthly_rows.append(row_data)

    monthly_df = pd.DataFrame(monthly_rows).reindex(columns=df.columns)
    return pd.concat([df, monthly_df], ignore_index=True)


def parse_bob_source(uploaded_file) -> pd.DataFrame:
    """Dispatch on file extension: .csv is the standard Business on the
    Books export every hotel uses; .xlsx is Margaritaville's differently-
    formatted PMS export (SR + Forecast wired up so far — ROB needs a small
    additional tweak once that's confirmed)."""
    file_bytes = uploaded_file.read()
    if uploaded_file.name.lower().endswith(".xlsx"):
        return parse_margaritaville_source(file_bytes)
    return parse_csv(file_bytes)


# ── ROB Update ───────────────────────────────────────────────────────────────

ROB_SHEETS = ["wk one", "wk two", "wk three", "wk four", "wk five", "wk six"]


def find_secondary_col(ws, block_start):
    candidates = []
    for cell in ws[block_start]:
        if cell.column <= 5:
            continue
        if isinstance(cell.value, str) and "variance" in cell.value.strip().lower():
            candidates.append(cell.column)
    return min(candidates) if candidates else None


def build_rob_change_plan(df, ws, grp_npu_rev_override: dict = None):
    """grp_npu_rev_override: optional {(year, month): dollar_value} — when
    present for a given month, writes that literal value into the 'Group Not
    P/U rev' secondary-column cell instead of the standard count*ADR formula.
    Used for Margaritaville, whose source data doesn't include a reliable
    Not-P/U room count to build that formula from — instead the value is
    computed elsewhere as the difference between two comparable PMS exports
    (one including not-yet-picked-up group revenue, one excluding it)."""
    today = datetime.date.today()
    current_month = today.month
    current_year = today.year
    changes = []

    # E4 = as-of date
    changes.append({
        "row": 4, "col": 5, "label": "As-of date", "month": None,
        "new_value": today, "skip_reason": None,
    })

    for _, row in df.iterrows():
        date_str = str(row[0]).strip() if row[0] else ""
        kind, info = classify_row(date_str)
        if kind != "monthly":
            continue
        year, month = info
        prev_month = current_month - 1 if current_month > 1 else 12
        prev_year  = current_year if current_month > 1 else current_year - 1
        if year == prev_year and month == prev_month:
            pass  # allow previous month (final numbers come in on the 1st)
        elif year != current_year or month < current_month:
            continue

        month_index = month - 1
        block_start = 4 + 8 * month_index

        rev     = safe_float(row[5])
        rms     = safe_float(row[1])
        grp_pu  = safe_float(row[7])
        grp_npu = safe_float(row[8])
        grp_rvn = safe_float(row[9])

        grp_sold = (grp_pu or 0) + (grp_npu or 0) if grp_pu is not None and grp_npu is not None else None
        sec_col = find_secondary_col(ws, block_start)

        entries = [
            (block_start + 1, 5, "Revenue",        rev,       False),
            (block_start + 2, 5, "Room Nights",     rms,       False),
            (block_start + 4, 5, "Group Rms Sold",  grp_sold,  False),
            (block_start + 5, 5, "Group Rm Rev",    grp_rvn,   False),
        ]
        if sec_col:
            npu_row    = block_start + 4
            entries.append((npu_row, sec_col, "Group Not P/U rooms", grp_npu, False))

            override_val = grp_npu_rev_override.get((year, month)) if grp_npu_rev_override else None
            if override_val is not None:
                entries.append((npu_row + 1, sec_col, "Group Not P/U rev (computed)", override_val, False))
            else:
                from openpyxl.utils import get_column_letter
                sec_letter = get_column_letter(sec_col)
                adr_row    = block_start + 6
                npu_formula = f"={sec_letter}{npu_row}*E{adr_row}"
                entries.append((npu_row + 1, sec_col, "Group Not P/U rev (formula)", npu_formula, True))

        for r, c, label, val, is_formula_write in entries:
            skip = None
            if r >= 100:
                skip = "row≥100"
            elif not is_formula_write and is_formula(ws.cell(r, c).value):
                skip = "formula"
            changes.append({"row": r, "col": c, "label": label, "month": month,
                             "new_value": val, "skip_reason": skip})

    return changes


def compute_grp_npu_rev_override(df, npu_compare_df):
    """Margaritaville ROB only: npu_compare_df is the same source format but
    from a second PMS export that includes not-yet-picked-up group revenue
    (df itself is the export that excludes it — confirmed the smaller of the
    two is used everywhere else). The difference per month is the dollar
    value for the "Group Not P/U rev" bright-green box on the ROB. Returns
    None if either input is missing (i.e. every hotel except Margaritaville)."""
    if npu_compare_df is None or df is None:
        return None

    def _monthly_col5(source_df):
        out = {}
        for _, row in source_df.iterrows():
            kind, info = classify_row(str(row[0]).strip() if row[0] else "")
            if kind == "monthly":
                out[info] = safe_float(row[5])
        return out

    main_sums    = _monthly_col5(df)
    compare_sums = _monthly_col5(npu_compare_df)
    return {
        key: compare_sums[key] - main_sums[key]
        for key in main_sums
        if key in compare_sums and main_sums[key] is not None and compare_sums[key] is not None
    }


def apply_rob_changes(wb, sheet_name, changes):
    ws = wb[sheet_name]
    for ch in changes:
        if ch["skip_reason"]:
            continue
        ws.cell(ch["row"], ch["col"]).value = ch["new_value"]


DONE_TAB_RGB = "FF00B050"  # green — set by color_tab_done() when a week is genuinely complete
DONE_TAB_HEX = "00B050"    # same green, without the alpha channel


def _is_done_color(rgb_value) -> bool:
    """True if rgb_value is our green 'done' marker. Matches on the trailing
    6 hex digits (case-insensitive) so it recognizes a tab colored directly in
    Excel's own Tab Color picker (which often stores 6-digit RGB with no alpha
    prefix) as well as ones this app set (8-digit ARGB) — a real week that a
    hotel had already filled in and marked green by hand was being treated as
    'not done' by a strict 8-char match, so the app kept re-picking it.
    """
    return isinstance(rgb_value, str) and rgb_value[-6:].upper() == DONE_TAB_HEX


def first_uncolored_sheet(wb, sheet_names):
    """Return the first ROB week tab that is neither marked done (our green)
    nor already holding real data in this month's block.

    Color alone isn't enough in either direction: a master template can carry
    its own unrelated baked-in tab color on a sheet that's never been touched
    (so "any color = done" produces false positives), while a real, already-
    filled week can fail to carry our exact green (so "only our green = done"
    produces false negatives — confirmed on a real hotel: week one already had
    this month's Revenue/Room Nights filled in but got silently re-picked and
    overwritten because its tab color didn't match). Checking actual data in
    the cells this month's update is about to write closes that gap.
    """
    today = datetime.date.today()
    block_start = 4 + 8 * (today.month - 1)
    for name in sheet_names:
        ws = wb[name]
        tc = ws.sheet_properties.tabColor
        if tc is not None and _is_done_color(getattr(tc, "rgb", None)):
            continue
        rev = ws.cell(block_start + 1, 5).value
        rms = ws.cell(block_start + 2, 5).value
        if isinstance(rev, (int, float)) or isinstance(rms, (int, float)):
            continue
        return name
    return sheet_names[-1]  # fallback: last sheet


def first_unhighlighted_forecast_sheet(wb, sheet_names):
    """Return the first Forecast week tab that is neither marked done (our
    green) nor already holding real OTB Rooms Sold data.

    Tab color alone can't be trusted in either direction here: hotels color-
    code Forecast tabs by hand with whatever color they like, but the SAME
    color can also be a stray artifact baked into a never-used master
    template — confirmed on real files: Hampton's untouched master has ALL 9
    week tabs pre-colored, and Provincetown Brass's untouched master has
    FCST-WK1 pre-colored magenta despite being completely blank. Checking the
    OTB Rooms Sold row for actual filled-in numbers is the reliable signal;
    our own exact green is kept as a secondary check for weeks we marked done
    ourselves but that ended up with no literal numbers written (e.g. every
    cell in that row happened to be formula-protected).
    """
    for name in sheet_names:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        tc = ws.sheet_properties.tabColor
        if tc is not None and _is_done_color(getattr(tc, "rgb", None)):
            continue
        rows = locate_forecast_rows(ws)
        if not rows:
            return name  # can't verify — treat as available rather than guess
        otb_row = rows["otb_rooms_row"]
        has_data = any(isinstance(ws.cell(otb_row, c).value, (int, float)) for c in range(2, 10))
        if has_data:
            continue
        return name
    return sheet_names[-1]  # fallback: last sheet


def first_undone_strategy_sheet(wb, sheet_names):
    """Return the first Strategy Report week tab that is neither marked done
    (our green) nor already holding real OTB TY Trans data.

    Same reasoning as first_unhighlighted_forecast_sheet: tab color alone
    can't be trusted — confirmed on real files that the identical purple
    (FF9900FF) marks a genuinely completed Surfside week AND sits untouched
    on Wolfboro's never-used master template. Check actual filled data as the
    reliable signal; our own exact green is a secondary check for weeks we
    marked done ourselves but that ended up with no literal numbers written.
    """
    for name in sheet_names:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        tc = ws.sheet_properties.tabColor
        if tc is not None and _is_done_color(getattr(tc, "rgb", None)):
            continue
        col_map = detect_strategy_columns(ws)
        otb_col = col_map.get("otb_trans")
        if not otb_col:
            return name  # can't verify — treat as available rather than guess
        has_data = any(isinstance(ws.cell(r, otb_col).value, (int, float)) for r in range(5, 15))
        if has_data:
            continue
        return name
    return sheet_names[-1]  # fallback: last sheet


def color_tab_done(wb, sheet_name):
    """Mark a sheet tab green to indicate it has been completed."""
    from openpyxl.styles.colors import Color
    wb[sheet_name].sheet_properties.tabColor = Color(rgb=DONE_TAB_RGB)


def clear_tab_colors(wb, sheet_names):
    """Reset tab color to none for every listed sheet. New-month setup copies
    the master/prior file, which can carry over a stale 'done' tab color —
    without this, a week that hasn't been touched yet can look completed and
    get skipped by the 'first uncolored sheet' auto-detect."""
    for name in sheet_names:
        if name in wb.sheetnames:
            wb[name].sheet_properties.tabColor = None


def strip_tables(wb):
    """Remove Excel table definitions to prevent openpyxl save corruption."""
    for ws in wb.worksheets:
        ws.tables.clear()


# ── Strategy Report ───────────────────────────────────────────────────────────

STRATEGY_SHEETS = ["WKONE", "WKTWO", "WKTHREE", "WKFOUR", "WKFIVE"]
FORECAST_SHEETS = ["FCST-WK1", "FCST-WK2", "FCST-WK3", "FCST-WK4",
                   "FCST-WK5", "FCST-WK6", "FCST-WK7", "FCST-WK8", "FCST-WK9"]

# CSV column index for each field (0-based) — source data never changes
STRATEGY_CSV_COLS = {
    "otb_trans":    (15, "OTB TY Trans (Indiv Count)"),
    "grp_pu_ty":    ( 7, "GRP PU TY"),
    "grp_npu_ty":   ( 8, "GRP N/PU TY"),
    "ooo_rms":      ( 4, "OOO RMS"),
    "trans_rev_ty": (16, "Trans Rev TY"),
    "grp_rev_ty":   ( 9, "Grp Rev TY"),
}

# Each field: list of (row3_keyword, row4_keyword) pairs to try in order.
# Match = both keywords found (case-insensitive) in their respective rows of that column.
# A None keyword means "don't check that row."
# "!WORD" suffix(es) on a keyword mean the combined headers must NOT contain WORD.
STRATEGY_FIELD_PATTERNS = {
    # ── TY columns (written from CSV) ──────────────────────────────────────────
    "otb_trans":       [("OTB TY", "TRANS"),            ("TRANS!LY", "SOLD!LY")],
    "grp_pu_ty":       [("GRP PU", "TY!LY"),            ("GROUP!LY", "SOLD!LY")],
    "grp_npu_ty":      [("GRP N/PU", "TY!LY"),          ("GRP RMS", "N/PU"),       ("N/PU!LY", None)],
    "ooo_rms":         [("OOO", None)],
    "trans_rev_ty":    [("TY TRANS", "REV"),             ("TRAN!LY", "REV TY")],
    "grp_rev_ty":      [("GRP TY", "REV"),               ("GRP!LY!N/PU", "REV TY")],
    "otb_lst_wk":      [("OTB", "LST WEK"),               ("OTB", "LST WK"),         ("OTB", "LAST WK"), ("OTB LST", None)],
    # ── LY columns (written from last year's SR) ───────────────────────────────
    "otb_ly_trans":    [("LY", "TRAN"),                  ("OTB LY", "TRANS"),       ("TRANS!TY", "SOLD!TY"), ("LY", "TRANS!TY")],
    "grp_pu_ly":       [("LY", "GRP"),                   ("GRP PU", "LY"),          ("GROUP!TY", "LY"),        ("GRP PU LY", None)],
    "grp_npu_ly":      [("GRP N/PU", "LY"),              ("N/PU LY", None),         ("GRP RMS", "LY")],
    "trans_rev_ly":    [("LY TRANS", "REV"),             ("TRAN!TY", "REV LY"),     ("LY", "TRANS REV")],
    "grp_rev_ly":      [("GRP LY", "REV"),               ("GRP!TY!N/PU", "REV LY")],
    "grp_npu_rev_ly":  [("GRP N/PU", "REV LY"),         ("N/PU LY", "REV"),        ("N/PU", "REV LY")],
}

# Maps LY destination field → TY source field in last year's SR
LY_FROM_TY = {
    "otb_ly_trans":   "otb_trans",
    "grp_pu_ly":      "grp_pu_ty",
    "grp_npu_ly":     "grp_npu_ty",
    "trans_rev_ly":   "trans_rev_ty",
    "grp_rev_ly":     "grp_rev_ty",
    "grp_npu_rev_ly": "grp_npu_ty",  # source is GRP N/PU TY
}

def _kw_matches(cell_val, keyword, r3_val, r4_val):
    """Check if keyword matches cell_val.
    Supports !WORD suffixes — the combined headers must NOT contain those words.
    e.g. 'TRAN!LY!ADR' matches if cell contains 'TRAN' and neither header contains 'LY' or 'ADR'.
    """
    parts = keyword.split("!")
    kw = parts[0].strip()
    excludes = [p.strip().upper() for p in parts[1:]]
    if kw and kw.upper() not in str(cell_val or "").upper():
        return False
    combined = (str(r3_val or "") + " " + str(r4_val or "")).upper()
    for excl in excludes:
        if excl in combined:
            return False
    return True


def detect_strategy_columns(ws):
    """Scan rows 3+4 of THIS sheet and return {field_key: col_index} for each
    field. Every sheet is re-scanned independently — never assume two sheets
    (even in the same workbook/hotel) share column positions. Week-1 vs
    week-2+ tabs can differ (e.g. an extra pickup-tracking column shifts
    everything after it), so a value pinned from one sheet can silently be
    wrong on another.
    """
    max_col = ws.max_column
    # Build lookup: col → (r3_text, r4_text)
    headers = {}
    for c in range(1, max_col + 1):
        headers[c] = (
            str(ws.cell(3, c).value or "").strip(),
            str(ws.cell(4, c).value or "").strip(),
        )

    col_map = {}
    for field, patterns in STRATEGY_FIELD_PATTERNS.items():
        found = None
        for r3_kw, r4_kw in patterns:
            for c, (r3v, r4v) in headers.items():
                r3_ok = r3_kw is None or _kw_matches(r3v, r3_kw, r3v, r4v)
                r4_ok = r4_kw is None or _kw_matches(r4v, r4_kw, r3v, r4v)
                if r3_ok and r4_ok and (r3_kw or r4_kw):
                    found = c
                    break
            if found:
                break
        if found:
            col_map[field] = found
        else:
            col_map[field] = None  # will surface as a warning, not a crash

    return col_map


def detect_date_column(ws):
    """Find the column whose data rows (5+) contain the earliest daily dates —
    i.e. the column that maps to each row's actual calendar date.
    Scans cols 1-10 only (dates are always on the left side).

    A sheet commonly has a Last Year date column right next to the This Year
    one (e.g. col 1 = LY dates, col 3 = TY dates), both starting on the 1st
    of the same month in different years — under a pure "most consecutive,
    prefer earliest day-of-month" score they tie exactly, and confirmed on a
    real file (Anchor In) the tie silently kept the LY column, so the whole
    date-to-row map was built a year off and almost nothing matched the CSV.
    Breaking ties by which column's dates start closest to today reliably
    picks the current/forward-looking column instead.

    A freshly-set-up month's TY column is typically a literal anchor date
    followed by '=prevRow+1' formula rows (build_date_row_map already
    assumes and extrapolates this pattern) — counting only literal
    isinstance(date) values undercounts it, so a fully-literal LY column
    can win outright before the tie-break above even applies. Confirmed
    real case: Provincetown Harbor Hotel's SR — the LY (2025) column was
    fully literal while TY (2026) was anchor+formulas, so TY never reached
    the 3-date minimum and was skipped as a candidate entirely, silently
    picking LY and mapping every row a year behind the CSV's dates.
    Counting formula rows that follow a literal anchor as continuing the
    sequence (without evaluating them) fixes this the same way
    build_date_row_map already trusts that pattern.
    """
    today = datetime.date.today()
    best_col, best_consecutive, best_proximity = 3, -1, None  # fallback to col 3
    for c in range(1, 11):
        anchor_date = None
        count = 0
        for r in range(5, min(ws.max_row + 1, 15)):
            v = ws.cell(r, c).value
            if isinstance(v, datetime.datetime):
                d = v.date()
            elif isinstance(v, datetime.date):
                d = v
            elif anchor_date is not None and isinstance(v, str) and v.startswith("="):
                count += 1  # formula row — trust it continues the sequence
                continue
            else:
                continue
            if anchor_date is None:
                anchor_date = d
            count += 1
        if count < 3 or anchor_date is None:
            continue
        proximity = abs((anchor_date - today).days)
        if count > best_consecutive or (
            count == best_consecutive and (best_proximity is None or proximity < best_proximity)
        ):
            best_consecutive = count
            best_proximity = proximity
            best_col = c
    return best_col


def detect_comp_set_columns(ws, col_map):
    """Find the comp set chart's TY (far-left) and LY (far-right) columns.
    The chart sits between Restrictions and TY TRANS REV.
    LY col  = last non-empty text header to the LEFT of TY TRANS REV.
    TY col  = first non-empty text header to the RIGHT of Restrictions.
    Scans rows 2-6 for headers (tolerates slight layout shifts).
    Returns (ty_col, ly_col) or (None, None) if not found.
    """
    trans_rev_col = col_map.get("trans_rev_ty")
    if not trans_rev_col:
        return None, None

    restrict_col = find_restrictions_col(ws, upto_col=trans_rev_col - 1)

    left_bound  = (restrict_col + 1) if restrict_col else max(1, trans_rev_col - 30)
    right_bound = trans_rev_col - 1

    def _has_text_header(c):
        for scan_r in range(2, 7):
            v = str(ws.cell(scan_r, c).value or "").strip()
            if v and not v.replace(".", "").isdigit():
                return True
        return False

    # LY col: scan left from just before TY TRANS REV
    ly_col = None
    for c in range(right_bound, left_bound - 1, -1):
        if _has_text_header(c):
            ly_col = c
            break

    # TY col: scan right from just after Restrictions
    ty_col = None
    for c in range(left_bound, right_bound + 1):
        if _has_text_header(c):
            ty_col = c
            break

    if ty_col == ly_col:
        return None, None  # same column — chart not found

    return ty_col, ly_col


def get_ly_sr_data(service, hotel_id, hotel_name, current_month, sheet_name):
    """Fetch LY data from last year's SR (same month, same week tab).
    Returns {date: {ly_field: value}} where dates are mapped to this year.
    Also returns (comp_ty_col_in_ly_ws, comp_ly_col_in_current_ws_placeholder)
    via a separate 'comp_set' key: {this_year_date: value}.
    """
    ly_month = current_month.replace(year=current_month.year - 1)
    result, err = resolve_drive_workbook(service, hotel_id, hotel_name,
                                         "Strategy Report", month_date=ly_month)
    if err or not result:
        return {}

    file_id, _ = result
    try:
        file_bytes = drive_download(service, file_id)
        ly_wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    except Exception:
        return {}

    if sheet_name not in ly_wb.sheetnames:
        return {}

    ly_ws = ly_wb[sheet_name]
    ly_col_map  = detect_strategy_columns(ly_ws)
    ly_date_col = detect_date_column(ly_ws)

    # Build date→row for last year's sheet
    ly_date_row = {}
    for r in range(5, ly_ws.max_row + 1):
        v = ly_ws.cell(r, ly_date_col).value
        if isinstance(v, datetime.datetime):
            ly_date_row[v.date()] = r
        elif isinstance(v, datetime.date):
            ly_date_row[v] = r

    # Comp set: far-left TY col in LY sheet
    comp_ty_col, _ = detect_comp_set_columns(ly_ws, ly_col_map)

    out = {}  # {this_year_date: {field: value}}
    for ly_date, r in ly_date_row.items():
        this_year_date = ly_date.replace(year=ly_date.year + 1)
        row_data = {}

        # Pull each TY source field from last year
        for ly_dest_field, ty_src_field in LY_FROM_TY.items():
            src_col = ly_col_map.get(ty_src_field)
            if src_col:
                v = ly_ws.cell(r, src_col).value
                if v is not None and not is_formula(v):
                    row_data[ly_dest_field] = safe_float(v)

        # Comp set TY value (far-left hotel col in LY sheet) — keep text as-is (e.g. "Sold out", "LOS2")
        if comp_ty_col:
            v = ly_ws.cell(r, comp_ty_col).value
            if v is not None and not is_formula(v):
                row_data["comp_set_ly"] = v

        if row_data:
            out[this_year_date] = row_data

    return out


def build_date_row_map(wb):
    """Build {date: row_number} from WKONE using auto-detected date column.
    When subsequent rows contain formulas (=C5+1 style), extrapolates from the
    first real date so the full year is mapped correctly.
    """
    ws = wb["WKONE"]
    date_col = detect_date_column(ws)
    mapping = {}
    anchor_date = None
    anchor_row  = None
    for row_num in range(5, ws.max_row + 1):
        val = ws.cell(row_num, date_col).value
        if isinstance(val, datetime.datetime):
            d = val.date()
        elif isinstance(val, datetime.date):
            d = val
        elif anchor_date and isinstance(val, str) and val.startswith("="):
            # Formula row — extrapolate from anchor
            offset = row_num - anchor_row
            d = anchor_date + datetime.timedelta(days=offset)
        else:
            continue
        if anchor_date is None:
            anchor_date = d
            anchor_row  = row_num
        mapping[d] = row_num
    return mapping


def find_otb_date_cell(ws):
    """Return (row, col) of the as-of date cell — one row above the OTB/TRANS header.
    Tries both Plymouth-style (OTB TY / TRANS) and Long Beach-style (TRANS / SOLD).
    """
    col_map = detect_strategy_columns(ws)
    otb_col = col_map.get("otb_trans")
    if otb_col:
        # Find which of rows 3 or 4 has the header, then go one above
        for r in range(2, 6):
            v = str(ws.cell(r, otb_col).value or "").strip().upper()
            if "OTB" in v or "TRANS" in v or "SOLD" in v:
                return r - 1, otb_col
    return 2, 4  # fallback


def _extract_otb_trans_by_date(wb, sheet_name, from_date):
    """Read OTB TY Trans keyed by date from an in-memory workbook.
    Uses build_date_row_map for date→row resolution so formula-based date
    cells (=C5+1 style) are handled via anchor extrapolation.
    Only returns rows where date >= from_date.
    """
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    col_map = detect_strategy_columns(ws)
    otb_col = col_map.get("otb_trans")
    if not otb_col:
        return {}
    date_row_map = build_date_row_map(wb)  # reads WKONE, handles formula dates
    out = {}
    for d, r in date_row_map.items():
        if d < from_date:
            continue
        val = ws.cell(r, otb_col).value
        if val is not None and not is_formula(val):
            out[d] = safe_float(val)
    return out


def _extract_ly_data_from_wb(ly_wb, sheet_name):
    """Read all LY source fields + comp set TY col from an in-memory workbook.
    Returns {this_year_date: {field: value}} (dates shifted +1 year).
    """
    if sheet_name not in ly_wb.sheetnames:
        return {}
    ws = ly_wb[sheet_name]
    col_map  = detect_strategy_columns(ws)
    date_col = detect_date_column(ws)
    comp_ty_col, _ = detect_comp_set_columns(ws, col_map)

    out = {}
    for r in range(5, ws.max_row + 1):
        v = ws.cell(r, date_col).value
        if isinstance(v, datetime.datetime): d = v.date()
        elif isinstance(v, datetime.date):   d = v
        else: continue
        # Shift +1 day to match DOW: DOW shifts by 1 in a non-leap year, so
        # LY July 2 (Wed) should fill TY July 1 (Wed), not LY July 1 (Tue).
        this_year = d.replace(year=d.year + 1) - datetime.timedelta(days=1)
        row_data = {}
        for ly_dest, ty_src in LY_FROM_TY.items():
            src_col = col_map.get(ty_src)
            if src_col:
                val = ws.cell(r, src_col).value
                if val is not None and not is_formula(val):
                    row_data[ly_dest] = safe_float(val)
        if comp_ty_col:
            val = ws.cell(r, comp_ty_col).value
            if val is not None and not is_formula(val):
                row_data["comp_set_ly"] = val  # preserve text values like "Sold out", "LOS2"
        if row_data:
            out[this_year] = row_data
    return out


def build_strategy_change_plan(df, wb, sheet_name, prev_month_wb=None, ly_wb=None,
                               scope_start=None, scope_end=None):
    """Build strategy changes.
    prev_month_wb: in-memory previous month's SR workbook (for OTB Lst Wek on WKONE)
    ly_wb:         in-memory last year's SR workbook (for all LY columns, every week)
    """
    today = datetime.date.today()
    if scope_start is None:
        scope_start = today.replace(day=1)
    if scope_end is None:
        scope_end = datetime.date(today.year + 1, 12, 31)

    date_row_map = build_date_row_map(wb)
    ws = wb[sheet_name]

    # Detect actual column positions from headers — no guessing
    col_map = detect_strategy_columns(ws)
    ly_only_fields = {"otb_lst_wk", "otb_ly_trans", "grp_pu_ly", "grp_npu_ly",
                      "trans_rev_ly", "grp_rev_ly", "grp_npu_rev_ly"}
    missing = [f for f, c in col_map.items() if c is None and f not in ly_only_fields]
    if missing:
        st.warning(f"Strategy: could not locate columns for: {', '.join(missing)}")

    # Comp set columns in current sheet
    comp_ty_col_cur, comp_ly_col_cur = detect_comp_set_columns(ws, col_map)

    # OTB Lst Wek — WKONE only, from previous month's SR (already in memory)
    prev_otb_map = {}
    src_sheet = None
    if sheet_name == "WKONE" and col_map.get("otb_lst_wk") and prev_month_wb:
        # Use last FILLED tab (opposite of first_undone) — checked by actual
        # OTB data, not tab color (see first_undone_strategy_sheet: the same
        # color can mark a genuinely completed week OR sit untouched on a
        # never-used master template, so color alone can't tell "filled").
        last_filled = None
        for s in STRATEGY_SHEETS:
            if s not in prev_month_wb.sheetnames:
                continue
            pcol_map = detect_strategy_columns(prev_month_wb[s])
            potb_col = pcol_map.get("otb_trans")
            if potb_col and any(isinstance(prev_month_wb[s].cell(r, potb_col).value, (int, float)) for r in range(5, 15)):
                last_filled = s
        src_sheet = last_filled or (STRATEGY_SHEETS[-1] if STRATEGY_SHEETS[-1] in prev_month_wb.sheetnames else None)
        if src_sheet:
            prev_otb_map = _extract_otb_trans_by_date(prev_month_wb, src_sheet, scope_start)

    # LY data — every week, from last year's same month/week tab (already in memory)
    ly_data = {}
    if ly_wb:
        ly_data = _extract_ly_data_from_wb(ly_wb, sheet_name)

    changes = []

    # Today's date above the OTB TY TRANS header
    date_row, date_col = find_otb_date_cell(ws)
    if date_row >= 1:
        changes.append({
            "date": today, "row": date_row, "col": date_col,
            "label": "As-of date", "new_value": today,
            "skip_reason": "formula" if is_formula(ws.cell(date_row, date_col).value) else None,
        })

    # As-of date above OTB Lst Wek header (row above row 3, i.e. row 2)
    lst_wk_col = col_map.get("otb_lst_wk")
    if lst_wk_col and prev_month_wb and src_sheet:
        prev_ws = prev_month_wb[src_sheet]
        prev_date_row, _ = find_otb_date_cell(prev_ws)
        src_date = prev_ws.cell(prev_date_row, date_col).value if prev_date_row >= 1 else None
        if src_date is None:
            # fallback: look in row above the lst_wk header in source
            for hr in range(2, 6):
                if str(prev_ws.cell(hr, lst_wk_col).value or "").strip():
                    src_date = prev_ws.cell(hr - 1, lst_wk_col).value
                    break
        if src_date:
            hdr_row = next((r for r in range(2, 6) if str(ws.cell(r, lst_wk_col).value or "").strip()), 3)
            label_row = hdr_row - 1
            if label_row >= 1:
                changes.append({
                    "date": None, "row": label_row, "col": lst_wk_col,
                    "label": "OTB Lst Wek as-of date",
                    "new_value": src_date,
                    "skip_reason": "formula" if is_formula(ws.cell(label_row, lst_wk_col).value) else None,
                })

    # As-of date above OTB LY TRANS header
    ly_trans_col = col_map.get("otb_ly_trans")
    if ly_trans_col and ly_wb and sheet_name in ly_wb.sheetnames:
        ly_ws_src = ly_wb[sheet_name]
        ly_date_row, _ = find_otb_date_cell(ly_ws_src)
        ly_src_date = ly_ws_src.cell(ly_date_row, date_col).value if ly_date_row >= 1 else None
        if ly_src_date:
            if isinstance(ly_src_date, datetime.datetime):
                ly_src_date = ly_src_date.date()
            hdr_row = next((r for r in range(2, 6) if str(ws.cell(r, ly_trans_col).value or "").strip()), 3)
            label_row = hdr_row - 1
            if label_row >= 1:
                changes.append({
                    "date": None, "row": label_row, "col": ly_trans_col,
                    "label": "OTB LY Trans as-of date",
                    "new_value": ly_src_date,
                    "skip_reason": None,
                })

    # ── CSV-sourced TY columns (only when BOB uploaded) ──────────────────────
    for _, row in (df.iterrows() if df is not None else []):
        date_str = str(row[0]).strip() if row[0] else ""
        kind, info = classify_row(date_str)
        if kind != "daily":
            continue
        d = info
        if d < scope_start or d > scope_end:
            continue
        if d not in date_row_map:
            continue
        excel_row = date_row_map[d]
        for field, (csv_col, label) in STRATEGY_CSV_COLS.items():
            excel_col = col_map.get(field)
            if excel_col is None:
                continue
            val = safe_float(row[csv_col])
            # BOB data is always authoritative — write even if cell has a template formula
            changes.append({
                "date": d, "row": excel_row, "col": excel_col,
                "label": label, "new_value": val, "skip_reason": None,
            })

    # ── Drive-sourced columns (run every time, no CSV needed) ─────────────────
    all_dates = set(date_row_map.keys()) & (set(prev_otb_map) | set(ly_data))
    for d in sorted(all_dates):
        if d < scope_start or d > scope_end:
            continue
        excel_row = date_row_map[d]

        # OTB Lst Wek — WKONE only
        lst_wk_col = col_map.get("otb_lst_wk")
        if lst_wk_col and d in prev_otb_map:
            skip = "formula" if is_formula(ws.cell(excel_row, lst_wk_col).value) else None
            changes.append({
                "date": d, "row": excel_row, "col": lst_wk_col,
                "label": "OTB Lst Wek", "new_value": prev_otb_map[d], "skip_reason": skip,
            })

        # LY columns
        if d in ly_data:
            row_ly = ly_data[d]
            for ly_field, ly_label in [
                ("otb_ly_trans",   "OTB LY Trans"),
                ("grp_pu_ly",      "GRP PU LY"),
                ("grp_npu_ly",     "GRP N/PU LY"),
                ("trans_rev_ly",   "LY Trans Rev"),
                ("grp_rev_ly",     "GRP LY Rev"),
                ("grp_npu_rev_ly", "GRP N/PU LY Rev"),
            ]:
                dest_col = col_map.get(ly_field)
                if dest_col and ly_field in row_ly:
                    changes.append({
                        "date": d, "row": excel_row, "col": dest_col,
                        "label": ly_label, "new_value": row_ly[ly_field], "skip_reason": None,
                    })

            # Comp set far-right (LY) ← far-left (TY) from last year
            if comp_ly_col_cur and "comp_set_ly" in row_ly:
                changes.append({
                    "date": d, "row": excel_row, "col": comp_ly_col_cur,
                    "label": "Comp Set LY", "new_value": row_ly["comp_set_ly"], "skip_reason": None,
                })

    # ── Blank LY cells with no confirmed LY data — don't leave stale leftovers
    # from whatever the template last held (e.g. a day with no LY rates). Only
    # runs when last year's SR actually loaded, so a missing/failed ly_wb never
    # wipes cells out of ignorance.
    if ly_wb:
        ly_field_labels = [
            ("otb_ly_trans",   "OTB LY Trans"),
            ("grp_pu_ly",      "GRP PU LY"),
            ("grp_npu_ly",     "GRP N/PU LY"),
            ("trans_rev_ly",   "LY Trans Rev"),
            ("grp_rev_ly",     "GRP LY Rev"),
            ("grp_npu_rev_ly", "GRP N/PU LY Rev"),
        ]
        for d, excel_row in date_row_map.items():
            if d < scope_start or d > scope_end:
                continue
            row_ly = ly_data.get(d, {})
            for ly_field, ly_label in ly_field_labels:
                dest_col = col_map.get(ly_field)
                if not dest_col or ly_field in row_ly:
                    continue
                cur_val = ws.cell(excel_row, dest_col).value
                if cur_val is None or is_formula(cur_val):
                    continue
                changes.append({
                    "date": d, "row": excel_row, "col": dest_col,
                    "label": f"{ly_label} (no LY data — cleared)",
                    "new_value": None, "skip_reason": None,
                })
            if comp_ly_col_cur and "comp_set_ly" not in row_ly:
                cur_val = ws.cell(excel_row, comp_ly_col_cur).value
                if cur_val is not None and not is_formula(cur_val):
                    changes.append({
                        "date": d, "row": excel_row, "col": comp_ly_col_cur,
                        "label": "Comp Set LY (no LY data — cleared)",
                        "new_value": None, "skip_reason": None,
                    })

    return changes


def apply_strategy_changes(wb, sheet_name, changes):
    ws = wb[sheet_name]
    for ch in changes:
        if ch["skip_reason"]:
            continue
        ws.cell(ch["row"], ch["col"]).value = ch["new_value"]


def parse_rate_csv(file_bytes: bytes) -> pd.DataFrame:
    df = pd.read_csv(io.BytesIO(file_bytes), dtype=str, encoding="utf-8-sig")
    df.columns = [c.strip() for c in df.columns]
    return df


def find_header_col(ws, keyword, header_rows=(2, 3, 4)):
    """Find the leftmost column whose concatenated header text (rows 2-4) contains keyword."""
    keyword = keyword.strip().lower()
    col_texts = {}
    for row_num in header_rows:
        for cell in ws[row_num]:
            col = cell.column
            if col not in col_texts:
                col_texts[col] = ""
            if isinstance(cell.value, str):
                col_texts[col] += cell.value.strip()
    matches = [col for col, text in col_texts.items() if keyword in text.lower()]
    return min(matches) if matches else None


def find_restrictions_col(ws, upto_col=None):
    """Find the 'Restrictions' header column, scanning rows 3-4.
    The header can be split across those two rows AND hyphenated for
    word-wrap (e.g. 'Restric'+'tions', or 'Restri-'+'ctions' — confirmed on
    a real hotel's sheet, ALLEGRIA/'Long Beach'). Stripping spaces and
    hyphens before matching handles both forms; a plain substring search on
    the raw concatenation misses the hyphenated one and returns nothing.

    Also confirmed a master template with the header simply misspelled —
    'Restictions' (missing the second 'r') on Hampton/Ashworth By The Sea, in
    a single cell, no split at all. Rather than chase every possible typo,
    match on 'REST' + 'TION' both present, which covers correct spelling,
    the hyphenated/split form, and this typo alike.
    """
    upto_col = upto_col or ws.max_column
    for c in range(1, upto_col + 1):
        combined_hdr = (str(ws.cell(3, c).value or "") + str(ws.cell(4, c).value or "")).upper().replace(" ", "").replace("-", "")
        if "REST" in combined_hdr and "TION" in combined_hdr:
            return c
    return None


def build_rates_change_plan(rate_df, wb, sheet_name):
    today = datetime.date.today()
    # include previous month — final numbers arrive on the 1st of the following month
    prev_month_start = (today.replace(day=1) - datetime.timedelta(days=1)).replace(day=1)
    scope_start = prev_month_start
    scope_end = datetime.date(today.year, 12, 31)

    date_row_map = build_date_row_map(wb)
    ws = wb[sheet_name]

    restric_col = find_restrictions_col(ws)
    hotel_col   = (restric_col + 1) if restric_col else None

    changes = []
    warnings = []
    if not restric_col:
        warnings.append("Could not find Restrictions column in sheet headers.")

    for _, row in rate_df.iterrows():
        date_str = str(row.get("Date", "")).strip()
        d = None
        for fmt in ("%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d"):
            try:
                d = datetime.datetime.strptime(date_str, fmt).date()
                break
            except ValueError:
                continue
        if d is None:
            continue
        if d < scope_start or d > scope_end:
            continue
        if d not in date_row_map:
            continue

        excel_row  = date_row_map[d]
        double_val = safe_float(row.get("Double", ""))
        mlos_val   = safe_float(row.get("Min Length of Stay", ""))

        if hotel_col:
            skip = "formula" if is_formula(ws.cell(excel_row, hotel_col).value) else None
            changes.append({"date": d, "row": excel_row, "col": hotel_col,
                            "label": "Hotel Rate", "new_value": double_val, "skip_reason": skip})
        if restric_col:
            skip = "formula" if is_formula(ws.cell(excel_row, restric_col).value) else None
            changes.append({"date": d, "row": excel_row, "col": restric_col,
                            "label": "Restrictions (MLOS)", "new_value": mlos_val, "skip_reason": skip})

    return changes, warnings


# ── Forecast helpers ─────────────────────────────────────────────────────────

DATE_FORMATS = [
    "%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%m/%d/%y", "%m-%d-%y",
    "%B %d, %Y", "%b %d, %Y", "%B %d %Y", "%b %d %Y",
    "%d-%b-%Y", "%d/%b/%Y",
]

def parse_any_date(val):
    """Parse a date from a datetime object, int serial, or string in many formats."""
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.date() if isinstance(val, datetime.datetime) else val
    if isinstance(val, (int, float)):
        # Excel serial date (days since 1900-01-01, with Lotus bug offset)
        return (datetime.date(1899, 12, 30) + datetime.timedelta(days=int(val)))
    if isinstance(val, str):
        val = val.strip()
        for fmt in DATE_FORMATS:
            try:
                return datetime.datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    return None


def locate_forecast_rows(ws):
    """Find the As-of date / OTB Rooms Sold / ADR OTB / actual Rooms Sold /
    actual Revenue rows by reading column-A titles, instead of assuming fixed
    row numbers. Confirmed on a real workbook that this drifts within a single
    file: WK1 has an extra 'Occupancy' row that WK4/WK8/WK9 don't, shifting
    every row below it by one — a hardcoded row 14/16/19 lands on a blank row
    or the wrong label on most week tabs.

    'Rooms Sold' and 'Revenue' each appear multiple times in this template, so
    rows are resolved in reading order relative to the unique 'ADR OTB' label:
    the first 'Rooms Sold' above it is the OTB (future) entry row, the first
    'Rooms Sold' below it is the actual entry row, and the first 'Revenue'
    below that is the actual revenue row.

    Returns None if the expected labels aren't all found (caller should warn
    and skip rather than guess a row number).
    """
    labels = []
    for r in range(1, 31):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip():
            labels.append((r, v.strip().lower()))

    def find_after(text, after_row=0):
        for r, v in labels:
            if r > after_row and text in v:
                return r
        return None

    dow_row            = find_after("day of week")
    otb_rooms_row      = find_after("rooms sold")
    adr_otb_row        = find_after("adr otb")
    actual_rooms_row   = find_after("rooms sold", adr_otb_row) if adr_otb_row else None
    actual_revenue_row = find_after("revenue", actual_rooms_row) if actual_rooms_row else None

    if not all([dow_row, otb_rooms_row, adr_otb_row, actual_rooms_row, actual_revenue_row]):
        return None

    return {
        "as_of_row":          dow_row - 1,
        "date_row":           dow_row + 1,
        "otb_rooms_row":      otb_rooms_row,
        "adr_otb_row":        adr_otb_row,
        "actual_rooms_row":   actual_rooms_row,
        "actual_revenue_row": actual_revenue_row,
    }


def build_forecast_date_col_map(ws, wb=None, date_row=4):
    """Return {date: col_index} from date_row. Falls back to WK1 for formula-only sheets."""
    month_start = parse_any_date(ws.cell(date_row, 2).value)

    # If this sheet's col B is a formula, find the start date from any WK sheet with a literal
    if month_start is None and wb is not None:
        for sname in wb.sheetnames:
            if "glance" in sname.lower():
                continue
            candidate = parse_any_date(wb[sname].cell(date_row, 2).value)
            if candidate is not None:
                month_start = candidate
                break

    if month_start is None:
        return {}

    col_map = {}
    col = 2
    while col <= ws.max_column:
        cell = ws.cell(date_row, col)
        if isinstance(cell.value, str) and "total" in cell.value.lower():
            break
        if cell.value is None and col > 2:
            break
        col_map[month_start + datetime.timedelta(days=col - 2)] = col
        col += 1
    return col_map


def row_is_filled(ws, r):
    """Return True if cols B-AF contain actual numbers OR cross-sheet formula refs."""
    for col in range(2, 33):
        v = ws.cell(r, col).value
        if isinstance(v, (int, float)):
            return True
        if isinstance(v, str) and v.startswith("='"):
            return True
    return False


def find_next_pickup_data_row(ws):
    """Find the next available row in the pick-up tracking chart.

    Strategy:
    1. Locate the 'Day of Week' section header (last one before 'Total Pick UP').
    2. Within the section, find the last row that has actual numbers in B-AF
       (cross-sheet formula refs look empty to openpyxl, so only direct numbers count).
    3. Scan forward from there for the first row with no numbers and no skip keyword
       — works for WK1/WK2 (back-to-back entries) and WK3+ (Pick UP rows between).
    """
    skip_keywords = {"pick", "day", "total", "forecast", "budget", "last year"}

    # Find Total Pick UP boundary
    total_row = None
    for r in range(40, 150):
        if "total pick" in str(ws.cell(r, 1).value or "").lower():
            total_row = r
            break
    search_end = total_row if total_row else 150

    # Find last 'Day of Week' header before the boundary
    section_start = None
    for r in range(40, search_end):
        if "day of week" in str(ws.cell(r, 1).value or "").lower():
            section_start = r
    if section_start is None:
        return None

    # Find the last row in the section that has actual numbers in B-AF
    last_filled = None
    for r in range(section_start + 2, search_end):
        if row_is_filled(ws, r):
            last_filled = r

    if last_filled is None:
        # Nothing filled yet — return first non-header row in section
        last_filled = section_start + 1

    # Scan forward from last_filled+1 for the first row with no numbers
    for r in range(last_filled + 1, search_end):
        a_val = str(ws.cell(r, 1).value or "").strip().lower()
        if any(k in a_val for k in skip_keywords):
            continue
        if not row_is_filled(ws, r):
            return r

    return None


def extract_rob_month_end_data(rob_wb, target_month):
    """Pull Budget and Last Year Room Nights + Revenue for target_month from any ROB sheet."""
    for sheet_name in rob_wb.sheetnames:
        ws = rob_wb[sheet_name]
        # Scan for the month block header row (col A = month abbreviation, e.g. "Jul")
        month_abbr = target_month.strftime("%b").lower()
        for r in range(1, ws.max_row + 1):
            cell_val = str(ws.cell(r, 1).value or "").strip().lower()
            if cell_val != month_abbr:
                continue
            # Found the block header row — find Budget and Last Year columns
            budget_col = ly_col = None
            for c in range(1, 25):
                hdr = str(ws.cell(r, c).value or "").strip()
                if not budget_col and "budget" in hdr.lower():
                    budget_col = c
                elif not ly_col and ("month end" in hdr.lower() or ("last" in hdr.lower() and "year" in hdr.lower())):
                    ly_col = c
                if budget_col and ly_col:
                    break
            # Also check row 4 (global header) if not found on block row
            if not budget_col or not ly_col:
                for c in range(1, 25):
                    hdr = str(ws.cell(4, c).value or "").strip()
                    if not budget_col and "budget" in hdr.lower():
                        budget_col = c
                    elif not ly_col and "month end" in hdr.lower():
                        ly_col = c
                    if budget_col and ly_col:
                        break
            if not budget_col or not ly_col:
                continue
            # Scan the next ~8 rows for Revenue and Room Nights labels in col A
            rev_row = rms_row = None
            for dr in range(1, 9):
                label = str(ws.cell(r + dr, 1).value or "").strip().lower()
                if "revenue" in label and rev_row is None:
                    rev_row = r + dr
                elif ("room" in label or "night" in label) and rms_row is None:
                    rms_row = r + dr
                if rev_row and rms_row:
                    break
            if not rev_row or not rms_row:
                continue
            return {
                "budget_rev": safe_float(ws.cell(rev_row, budget_col).value),
                "budget_rms": safe_float(ws.cell(rms_row, budget_col).value),
                "ly_rev":     safe_float(ws.cell(rev_row, ly_col).value),
                "ly_rms":     safe_float(ws.cell(rms_row, ly_col).value),
            }
    return None


def find_month_ending_forecast_cells(fcst_ws):
    """Return cell coords for Budget/LY Room Nts and Revenue in the Month Ending Forecast table."""
    for r in range(1, fcst_ws.max_row + 1):
        for c in range(1, fcst_ws.max_column + 1):
            if "month ending forecast" in str(fcst_ws.cell(r, c).value or "").lower():
                # Header found — scan next ~6 rows for Budget and Last Year rows
                # and the header row for column positions
                hdr_row = col_rms = col_rev = None
                budget_row = ly_row = None
                for dr in range(1, 8):
                    row_vals = {col: str(fcst_ws.cell(r + dr, col).value or "").strip()
                                for col in range(1, 8)}
                    row_text = " ".join(row_vals.values()).lower()
                    if "room" in row_text and hdr_row is None:
                        hdr_row = r + dr
                        for col, v in row_vals.items():
                            if "room" in v.lower():
                                col_rms = col
                            elif "revenue" in v.lower() or "rev" in v.lower():
                                col_rev = col
                    if "budget" in row_text and budget_row is None:
                        budget_row = r + dr
                    if "last year" in row_text and ly_row is None:
                        ly_row = r + dr
                if col_rms and col_rev and budget_row and ly_row:
                    return col_rms, col_rev, budget_row, ly_row
    return None, None, None, None


def build_forecast_change_plan(df, ws, rob_wb=None, is_wk1=False):
    """Build list of cell writes for the Forecast sheet."""
    today = datetime.date.today()
    yesterday = today - datetime.timedelta(days=1)
    month_start = today.replace(day=1)

    rows = locate_forecast_rows(ws)
    if not rows:
        return [], ["Could not read row titles (As-of date / Rooms Sold / ADR OTB / Revenue) from forecast sheet."]

    col_map = build_forecast_date_col_map(ws, ws.parent, date_row=rows["date_row"])
    if not col_map:
        return [], ["Could not read date row from forecast sheet."]

    changes = []
    warnings = []

    # As-of date (row directly above "Day of Week")
    changes.append({
        "label": "As-of date", "row": rows["as_of_row"], "col": 1,
        "new_value": today, "skip_reason": "formula" if is_formula(ws.cell(rows["as_of_row"], 1).value) else None,
    })

    # Build lookup from CSV: date -> row dict
    # Uses the same flexible date matching as ROB/SR (classify_row) — the BOB
    # CSV's date format (1-2 digit month/day, "/" or "-") wasn't matching the
    # old hyphen-only, zero-padded regex here, so Forecast silently got no data.
    daily_rows = {}
    for _, row in df.iterrows():
        date_str = str(row.iloc[0]).strip()
        kind, d = classify_row(date_str)
        if kind != "daily":
            continue
        daily_rows[d] = row

    for d, col in col_map.items():
        if d not in daily_rows:
            continue
        csv_row = daily_rows[d]
        rms = safe_float(csv_row.iloc[1])
        adr = safe_float(csv_row.iloc[6])
        rev = safe_float(csv_row.iloc[5])

        is_future = d >= today
        is_past   = d <= yesterday and d >= month_start

        # Rooms Sold (future / OTB)
        if is_future:
            skip = "formula" if is_formula(ws.cell(rows["otb_rooms_row"], col).value) else None
            changes.append({"label": f"Rooms Sold (future) {d}", "row": rows["otb_rooms_row"], "col": col,
                            "new_value": rms, "skip_reason": skip})
            # ADR OTB (future)
            skip = "formula" if is_formula(ws.cell(rows["adr_otb_row"], col).value) else None
            changes.append({"label": f"ADR OTB {d}", "row": rows["adr_otb_row"], "col": col,
                            "new_value": adr, "skip_reason": skip})

        # Rooms Sold (actuals)
        if is_past:
            skip = "formula" if is_formula(ws.cell(rows["actual_rooms_row"], col).value) else None
            changes.append({"label": f"Rooms Sold (actual) {d}", "row": rows["actual_rooms_row"], "col": col,
                            "new_value": rms, "skip_reason": skip})
            # Revenue (actuals)
            skip = "formula" if is_formula(ws.cell(rows["actual_revenue_row"], col).value) else None
            changes.append({"label": f"Revenue (actual) {d}", "row": rows["actual_revenue_row"], "col": col,
                            "new_value": rev, "skip_reason": skip})

    # Pick-up tracking row: write full month rooms sold to next available row
    target_row = find_next_pickup_data_row(ws)

    if target_row:
        changes.append({"label": "Pickup tracking: date", "row": target_row, "col": 1,
                        "new_value": today, "skip_reason": None})
        for d, col in col_map.items():
            if d not in daily_rows:
                continue
            rms = safe_float(daily_rows[d].iloc[1])
            skip = "formula" if is_formula(ws.cell(target_row, col).value) else None
            changes.append({"label": f"Pickup tracking: Rooms Sold {d}",
                            "row": target_row, "col": col,
                            "new_value": rms, "skip_reason": skip})
    else:
        warnings.append("No available row found in pick-up tracking chart.")

    # Month Ending Forecast table — only on WK1, only when ROB workbook provided
    if is_wk1 and rob_wb is not None:
        target_month = today.replace(day=1)
        rob_data = extract_rob_month_end_data(rob_wb, target_month)
        if rob_data:
            col_rms, col_rev, budget_row, ly_row = find_month_ending_forecast_cells(ws)
            if col_rms and col_rev and budget_row and ly_row:
                entries = [
                    (budget_row, col_rms, "Month End Forecast: Budget Room Nts", rob_data["budget_rms"]),
                    (budget_row, col_rev, "Month End Forecast: Budget Revenue",  rob_data["budget_rev"]),
                    (ly_row,     col_rms, "Month End Forecast: LY Room Nts",     rob_data["ly_rms"]),
                    (ly_row,     col_rev, "Month End Forecast: LY Revenue",       rob_data["ly_rev"]),
                ]
                for r, c, label, val in entries:
                    skip = "formula" if is_formula(ws.cell(r, c).value) else None
                    changes.append({"label": label, "row": r, "col": c,
                                    "new_value": val, "skip_reason": skip})
            else:
                warnings.append("Could not locate Month Ending Forecast table in forecast sheet.")
        else:
            warnings.append("Could not find month data in ROB workbook for Month Ending Forecast.")

    return changes, warnings


def build_next_month_forecast_plan(df, ws):
    """For weeks 3 & 4: write Rooms Sold and ADR OTB (rows located by title,
    not a fixed offset) for ALL dates in the next month (everything in the CSV
    beyond the current month). Also writes the As-of date and pick-up tracking row.
    """
    today = datetime.date.today()
    current_month_end = (today.replace(day=1) + datetime.timedelta(days=32)).replace(day=1) - datetime.timedelta(days=1)

    rows = locate_forecast_rows(ws)
    if not rows:
        return [], ["Could not read row titles (As-of date / Rooms Sold / ADR OTB / Revenue) from next-month forecast sheet."]

    col_map = build_forecast_date_col_map(ws, ws.parent, date_row=rows["date_row"])
    if not col_map:
        return [], ["Could not read date row from next-month forecast sheet."]

    changes = []
    warnings = []

    # As-of date (row directly above "Day of Week")
    changes.append({
        "label": "As-of date", "row": rows["as_of_row"], "col": 1,
        "new_value": today,
        "skip_reason": "formula" if is_formula(ws.cell(rows["as_of_row"], 1).value) else None,
    })

    daily_rows = {}
    for _, row in df.iterrows():
        date_str = str(row.iloc[0]).strip()
        kind, d = classify_row(date_str)
        if kind != "daily":
            continue
        daily_rows[d] = row

    for d, col in col_map.items():
        # Only next month dates (after current month end)
        if d <= current_month_end:
            continue
        if d not in daily_rows:
            continue
        csv_row = daily_rows[d]
        rms = safe_float(csv_row.iloc[1])
        adr = safe_float(csv_row.iloc[6])

        skip_rms = "formula" if is_formula(ws.cell(rows["otb_rooms_row"], col).value) else None
        skip_adr = "formula" if is_formula(ws.cell(rows["adr_otb_row"], col).value) else None
        changes.append({"label": f"Rooms Sold (future) {d}", "row": rows["otb_rooms_row"], "col": col, "new_value": rms, "skip_reason": skip_rms})
        changes.append({"label": f"ADR OTB {d}",             "row": rows["adr_otb_row"],   "col": col, "new_value": adr, "skip_reason": skip_adr})

    # Pick-up tracking row (same logic — full month of next month dates)
    target_row = find_next_pickup_data_row(ws)
    if target_row:
        changes.append({"label": "Pickup tracking: date", "row": target_row, "col": 1,
                        "new_value": today, "skip_reason": None})
        for d, col in col_map.items():
            if d not in daily_rows:
                continue
            rms = safe_float(daily_rows[d].iloc[1])
            skip = "formula" if is_formula(ws.cell(target_row, col).value) else None
            changes.append({"label": f"Pickup tracking: Rooms Sold {d}",
                            "row": target_row, "col": col, "new_value": rms, "skip_reason": skip})
    else:
        warnings.append("No available row found in next-month pick-up tracking chart.")

    return changes, warnings


def apply_forecast_changes(wb, sheet_name, changes):
    ws = wb[sheet_name]
    for ch in changes:
        if ch["skip_reason"]:
            continue
        ws.cell(ch["row"], ch["col"]).value = ch["new_value"]


# ── Google Drive ──────────────────────────────────────────────────────────────

MULTI_ID_PREFIX = "MULTI:"

# Hotels whose Drive folders had to be shared per-month/year directly (no
# common parent folder to share instead) get grouped by a known keyword
# instead of trying to generically parse the date prefix out of the folder
# name — that parsing kept breaking on real naming inconsistencies (mixed
# case, 2- vs 4-digit years, colon/period/no separator, even "Anchor In" vs
# "Hyannis Anchor In" for the same hotel). Matching on a fixed keyword that's
# guaranteed present in every one of that hotel's folder names is far more
# reliable. Add an entry here for each hotel using this sharing pattern.
KNOWN_MULTI_FOLDER_HOTELS = {
    "Hyannis Anchor In":     ["ANCHOR"],
    "Provincetown Surfside": ["SURFSIDE"],
    "Hotel 1620":            ["1620", "PLYMOUTH"],
    "Wolfeboro":             ["WOLF"],
    "Provincetown Harbor Hotel": ["HARBOR"],
}


def _strip_dedup_suffix(name):
    """Strip a trailing ' (1)', ' (2)', etc. — Drive's auto-added suffix when
    a folder name collides with an existing one (confirmed real case:
    Surfside's hotel folder is literally named 'SURFSIDE (1)' in Drive).
    Purely cosmetic for the dropdown; the real folder_id is unaffected.
    """
    return re.sub(r'\s*\(\d+\)\s*$', '', name).strip()


def _is_rev_reports_name(name):
    """True if a folder name denotes a 'Revenue Reports' folder. Confirmed
    real naming variants include the full phrase and the abbreviation
    'REV RPTS' (Hotel 1620's convention, e.g. 'G. JUL2018 REV RPTS HOTEL
    1620') — a plain 'REVENUE REPORTS' in name_upper substring check misses
    the abbreviated form entirely, which broke both hotel grouping and
    workbook resolution for that hotel.
    """
    return bool(re.search(r'REVENUE REPORTS|REV RPTS', name.upper()))


def _extract_hotel_name_from_rev_folder(name):
    """Strip date-ish prefixes and the phrase 'Revenue Reports' from a folder
    name, leaving just the hotel name. Confirmed real naming is inconsistent
    even for a single hotel (Hyannis Anchor In, sharing per-month/year folders
    since there's no common parent to share instead): 'A: APR2025 Revenue
    Reports Anchor In', 'A: JUN2025 Revenue Reports Hyannis Anchor In',
    'Jul2024 Revenue Reports Hyannis Anchor In', 'A: MAY25 Revenue Reports
    Hyannis Anchor In' — mixing a leading single-letter marker (with a colon,
    period, or dash — not just a period), 2- or 4-digit years, and even
    inconsistent inclusion of "Hyannis" in the hotel name itself.

    Only strips a leading single letter when followed by a clear separator
    (":", ".", "-") — a bare leading letter with no separator is the start of
    a month name (e.g. "Jul2024") and must be left alone, not eaten.
    """
    s = name.strip()
    s = re.sub(r'^[A-Za-z]\s*[\.\:\-]\s*', '', s)
    s = re.sub(r'^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\s*\d{2,4}\s+', '', s, flags=re.IGNORECASE)
    s = re.sub(r'^\d{4}\s+', '', s)
    s = re.sub(r'revenue\s*reports|rev\s+rpts', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\s+', ' ', s).strip()
    return s.strip(' -:.,')


@st.cache_data(ttl=300)
def get_hotels_from_drive():
    """Return list of (display_name, folder_id) for every top-level folder
    that contains a 'REVENUE REPORTS' subfolder — i.e. each hotel folder.
    Cached for 5 minutes so it doesn't hit Drive on every rerender.

    Some hotels are shared with the service account directly at the REVENUE
    REPORTS folder level, not its parent — this happens on Shared Drives
    where the person granting access can only share folders they themselves
    have permission on, and Drive permissions don't propagate upward to a
    parent. Some of those hotels additionally have a SEPARATE REVENUE
    REPORTS folder per year (confirmed real case: Hyannis Anchor In), each
    shared individually since there's no common parent to share instead —
    those get grouped into a single hotel entry by name, with folder_id set
    to 'MULTI:<id>,<id>,...' listing every year's folder. resolve_drive_workbook
    and _find_rev_reports_folder_for_year both know how to unpack this.
    """
    try:
        svc = get_drive_service()
        q = "mimeType = 'application/vnd.google-apps.folder' and trashed = false"
        result = svc.files().list(
            q=q, fields="files(id, name)", pageSize=200,
            supportsAllDrives=True, includeItemsFromAllDrives=True,
        ).execute()
        folders = result.get("files", [])

        hotels = []
        known_groups = {}  # hotel display name -> [folder_id, ...]
        rev_groups = []     # list of {"display": str, "ids": [folder_id, ...]} for unregistered hotels
        for folder in folders:
            name = folder["name"]
            name_upper = name.upper()

            # "Ancillary" folders are a different report category entirely
            # (ancillary revenue / front desk reports, not the ROB/SR/Forecast
            # per-month structure) — confirmed real case: "Ancillary
            # Provincetown Surfside" showing up as its own bogus hotel entry.
            # Exclude from grouping regardless of which hotel keyword matches.
            if "ANCILLARY" in name_upper:
                continue

            known_match = next((hn for hn, kws in KNOWN_MULTI_FOLDER_HOTELS.items()
                                 if any(kw in name_upper for kw in kws)), None)
            if known_match:
                known_groups.setdefault(known_match, []).append(folder["id"])
                continue

            if _is_rev_reports_name(name):
                extracted = _extract_hotel_name_from_rev_folder(name)
                if not extracted:
                    # No hotel name left after stripping the date prefix and
                    # "Revenue Reports" — an orphaned/mis-named folder with no
                    # usable hotel name (confirmed real case: "I. SEPT2021
                    # Revenue Reports", missing the hotel name entirely).
                    # Skip it rather than showing a meaningless entry.
                    continue
                norm = extracted.upper()
                # Merge by substring containment, not exact match — the same
                # hotel's own folders don't always spell its name the same way
                # (confirmed real case: some of Hyannis Anchor In's folders
                # say just "Anchor In", others "Hyannis Anchor In"). Keep the
                # longest/most complete variant seen as the display name.
                match = next((g for g in rev_groups
                              if norm in g["display"].upper() or g["display"].upper() in norm), None)
                if match:
                    match["ids"].append(folder["id"])
                    if len(extracted) > len(match["display"]):
                        match["display"] = extracted
                else:
                    rev_groups.append({"display": extracted, "ids": [folder["id"]]})
                continue
            if re.search(r'\b20\d{2}\b', name):
                continue
            child_q = ("'%s' in parents and trashed = false and "
                       "mimeType = 'application/vnd.google-apps.folder'") % folder["id"]
            children = svc.files().list(
                q=child_q, fields="files(name)", pageSize=20,
                supportsAllDrives=True, includeItemsFromAllDrives=True,
            ).execute()
            has_rev = any(_is_rev_reports_name(c["name"]) for c in children.get("files", []))
            if has_rev:
                hotels.append((_strip_dedup_suffix(name), folder["id"]))

        for hotel_name, ids in known_groups.items():
            hotels.append((_strip_dedup_suffix(hotel_name), ids[0] if len(ids) == 1 else MULTI_ID_PREFIX + ",".join(ids)))

        for info in rev_groups:
            display = _strip_dedup_suffix(info["display"])
            if len(info["ids"]) == 1:
                hotels.append((display, info["ids"][0]))
            else:
                hotels.append((display, MULTI_ID_PREFIX + ",".join(info["ids"])))

        return sorted(hotels, key=lambda x: x[0])
    except Exception:
        return []

WORKBOOK_TYPES = ["ROB", "Strategy Report", "Forecast"]

# Maps workbook type → partial filename keyword to search for in Drive
WORKBOOK_KEYWORDS = {
    "ROB":             "ROB",
    "Strategy Report": "STRATEGY",
    "Forecast":        "FORECAST",
}

CREDS_PATH = "credentials.json.json"
SCOPES     = ["https://www.googleapis.com/auth/drive"]


@st.cache_resource
def get_drive_service():
    # Production (Streamlit Cloud): credentials stored in st.secrets["google_credentials"]
    if "google_credentials" in st.secrets:
        import json
        creds = service_account.Credentials.from_service_account_info(
            dict(st.secrets["google_credentials"]), scopes=SCOPES
        )
    else:
        # Local dev: read from file
        creds = service_account.Credentials.from_service_account_file(CREDS_PATH, scopes=SCOPES)
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def drive_find_folder_by_keyword(service, keyword, parent_id=None):
    """Return the first folder whose name contains keyword (case-insensitive)."""
    q = "mimeType = 'application/vnd.google-apps.folder' and trashed = false"
    if parent_id:
        q += " and '%s' in parents" % parent_id
    result = service.files().list(
        q=q, fields="files(id, name)", pageSize=100,
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute()
    for f in result.get("files", []):
        if keyword.lower() in f["name"].lower():
            return f["id"], f["name"]
    return None, None


def _find_rev_reports_folder_for_year(service, hotel_id, year_kw):
    """Find the REVENUE REPORTS folder to use for a given year.
    Some hotels have ONE 'REVENUE REPORTS' folder for years directly (year/month
    subfolders inside it); others have a SEPARATE '<year> REVENUE REPORTS <hotel>'
    folder per year, sitting side by side. drive_find_folder_by_keyword only
    ever returns the first match, which silently picks the wrong year's folder
    for hotels using the per-year pattern. Prefer an exact year-name match;
    fall back to the first REVENUE REPORTS folder found.
    """
    # Hotel has a separate REVENUE REPORTS folder shared per year (each
    # individually shared, since there's no common parent to share instead).
    # Pick the one matching year_kw directly from the group.
    if hotel_id.startswith(MULTI_ID_PREFIX):
        candidate_ids = hotel_id[len(MULTI_ID_PREFIX):].split(",")
        candidates = []
        for cid in candidate_ids:
            try:
                info = service.files().get(fileId=cid, fields="name", supportsAllDrives=True).execute()
                candidates.append({"id": cid, "name": info["name"]})
            except Exception:
                continue
        year_match = next((f for f in candidates if year_kw in f["name"]), None)
        if year_match:
            return year_match["id"], year_match["name"]
        if candidates:
            return candidates[0]["id"], candidates[0]["name"]
        return None, None

    q = ("mimeType = 'application/vnd.google-apps.folder' and trashed = false "
         "and '%s' in parents") % hotel_id
    children = service.files().list(
        q=q, fields="files(id, name)", pageSize=100,
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute().get("files", [])
    candidates = [f for f in children if "revenue reports" in f["name"].lower()]
    year_match = next((f for f in candidates if year_kw in f["name"]), None)
    if year_match:
        return year_match["id"], year_match["name"]
    if candidates:
        return candidates[0]["id"], candidates[0]["name"]

    # Some hotels are shared directly at the REVENUE REPORTS folder level
    # (Shared Drive permissions don't propagate to a parent folder) — hotel_id
    # IS that folder already in that case, not its parent, so there's no
    # child to find. Check hotel_id's own name before giving up.
    try:
        self_info = service.files().get(
            fileId=hotel_id, fields="name", supportsAllDrives=True
        ).execute()
        if "revenue reports" in self_info.get("name", "").lower():
            return hotel_id, self_info["name"]
    except Exception:
        pass

    return None, None


def _find_or_create_month_folder_under_rev(service, rev_id, year_kw, month_kw, target_month, hotel_name):
    """Resolve the month folder for a new-month setup, handling both layouts:
    month folders directly inside the REVENUE REPORTS folder (common when
    there's one REVENUE REPORTS folder per year), or nested under a year
    subfolder. Creates the month folder if neither is found.
    """
    month_id, month_name = drive_find_folder_by_keyword(service, month_kw, parent_id=rev_id)
    if month_id:
        return month_id, month_name
    year_id, _ = drive_find_folder_by_keyword(service, year_kw, parent_id=rev_id)
    if year_id:
        return drive_find_or_create_month_folder(service, rev_id, year_id, target_month, hotel_name)
    return drive_find_or_create_month_folder(service, rev_id, rev_id, target_month, hotel_name)


def drive_find_file(service, keyword, parent_id):
    """Return (file_id, file_name) for first xlsx whose name contains keyword,
    excluding files whose name also contains 'copy' (to skip backup copies).

    Also matches native Google Sheets (mimeType 'application/vnd.google-apps.
    spreadsheet') — confirmed real case: Hotel 1620's Forecast workbook was
    created directly as a Google Sheet rather than uploaded as .xlsx, so an
    xlsx/xlsm-only filter silently missed it even though the name matched.
    drive_download handles exporting these to xlsx bytes on read.
    """
    q = ("'%s' in parents and trashed = false "
         "and (mimeType = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' "
         "or mimeType = 'application/vnd.ms-excel.sheet.macroenabled.12' "
         "or mimeType = 'application/vnd.google-apps.spreadsheet')") % parent_id
    result = service.files().list(
        q=q, fields="files(id, name)",
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute()
    files = result.get("files", [])
    # prefer non-copy files; fall back to copy if nothing else found
    files.sort(key=lambda f: (1 if "copy" in f["name"].lower() else 0))
    for f in files:
        name_lower = f["name"].lower()
        if keyword.lower() in name_lower and "master" not in name_lower:
            return f["id"], f["name"]
    return None, None


def drive_download(service, file_id) -> bytes:
    """Download a file's bytes. Native Google Sheets (created directly in
    Drive rather than uploaded as .xlsx — confirmed real case: Hotel 1620's
    Forecast workbook) can't be read via get_media like a normal blob file;
    they must be exported to xlsx format instead."""
    meta = service.files().get(fileId=file_id, fields="mimeType", supportsAllDrives=True).execute()
    buf = io.BytesIO()
    if meta.get("mimeType") == "application/vnd.google-apps.spreadsheet":
        req = service.files().export_media(
            fileId=file_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        req = service.files().get_media(fileId=file_id, supportsAllDrives=True)
    dl  = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    return buf.getvalue()


def drive_upload(service, file_id, file_bytes: bytes, file_name: str):
    """Overwrite an existing Drive file with new bytes."""
    buf   = io.BytesIO(file_bytes)
    media = MediaIoBaseUpload(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=True,
    )
    # supportsAllDrives is required here even though drive_download's
    # get_media() works without it — confirmed real case: writing back to a
    # Shared Drive file (Hyannis Anchor In) 404'd on update() alone, for
    # every workbook type, despite the same file_id having just been read
    # successfully moments earlier.
    service.files().update(fileId=file_id, media_body=media, supportsAllDrives=True).execute()


def drive_find_or_create_month_folder(service, rev_id: str, year_id: str, month_date: datetime.date, hotel_name: str):
    """Return (folder_id, folder_name) for the month folder under year_id.
    Folder name pattern: '[LETTER]: [MON][YEAR] REVENUE REPORTS [HOTEL_UPPER]'
    where LETTER = A-L for Jan-Dec.
    Creates the folder if it doesn't exist.
    """
    month_kw = month_date.strftime("%b%Y").upper()  # e.g. JUL2026

    # Try to find existing folder first
    q = ("mimeType='application/vnd.google-apps.folder' and trashed=false "
         "and '%s' in parents") % year_id
    result = service.files().list(q=q, fields="files(id,name)", pageSize=100).execute()
    for f in result.get("files", []):
        if month_kw in f["name"].upper():
            return f["id"], f["name"]

    # Not found — infer name from existing sibling folder naming pattern
    # Pattern: "[LETTER]: [MON][YEAR] REVENUE REPORTS [HOTEL]"
    letter = chr(ord("A") + month_date.month - 1)  # A=Jan, B=Feb, ..., L=Dec

    # Try to infer hotel suffix from existing month folders in this year folder
    hotel_suffix = hotel_name.upper()
    for f in result.get("files", []):
        name_upper = f["name"].upper()
        if "REVENUE REPORTS" in name_upper:
            # Extract everything after "REVENUE REPORTS "
            idx = name_upper.find("REVENUE REPORTS ")
            if idx != -1:
                hotel_suffix = f["name"][idx + len("REVENUE REPORTS "):].strip()
                break

    new_name = f"{letter}: {month_kw} REVENUE REPORTS {hotel_suffix}"
    folder_meta = {
        "name": new_name,
        "mimeType": "application/vnd.google-apps.folder",
        "parents": [year_id],
    }
    created = service.files().create(
        body=folder_meta, fields="id,name", supportsAllDrives=True,
    ).execute()
    return created["id"], created["name"]


def drive_copy_file(service, source_file_id: str, new_name: str, parent_folder_id: str):
    """Copy a Drive file to a new name in the given folder. Returns (new_file_id, new_name).
    supportsAllDrives=True is required when the destination is a shared drive —
    service accounts have no personal storage quota.
    """
    body = {"name": new_name, "parents": [parent_folder_id]}
    copied = service.files().copy(
        fileId=source_file_id, body=body, fields="id,name",
        supportsAllDrives=True,
    ).execute()
    return copied["id"], copied["name"]


def _hotel_search_scope_ids(service, hotel_id):
    """Return every folder id that could plausibly hold a hotel's MASTER
    template files: each of the hotel's own root candidate folder(s)
    (unwrapping a MULTI:<id>,<id>,... group) plus their direct children.

    Confirmed real case: master-file lookups used to search ALL of Drive
    with no folder scoping at all — Wolfeboro's 'Set Up New ROB' found and
    copied Hotel 1620's ROB master (Drive's unordered global search just
    happened to return it first), silently mislabeling the result 'JUL2026
    ROB PLYMOUTH.xlsx' inside Wolfeboro's own folder.
    """
    if hotel_id.startswith(MULTI_ID_PREFIX):
        root_ids = hotel_id[len(MULTI_ID_PREFIX):].split(",")
    else:
        root_ids = [hotel_id]

    scope_ids = list(root_ids)
    for rid in root_ids:
        q = ("mimeType = 'application/vnd.google-apps.folder' and trashed = false "
             "and '%s' in parents") % rid
        try:
            children = service.files().list(
                q=q, fields="files(id)", pageSize=100,
                supportsAllDrives=True, includeItemsFromAllDrives=True,
            ).execute().get("files", [])
        except Exception:
            children = []
        scope_ids.extend(c["id"] for c in children)
    return scope_ids


def find_rob_master(service, hotel_id: str):
    """Search the hotel's own Drive tree for the ROB master file."""
    scope_ids = _hotel_search_scope_ids(service, hotel_id)
    if not scope_ids:
        return None, "Could not resolve hotel folder to search."
    parent_clause = " or ".join("'%s' in parents" % sid for sid in scope_ids)
    q = ("trashed=false and (%s) "
         "and (mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' "
         "or mimeType='application/vnd.ms-excel.sheet.macroenabled.12') "
         "and name contains 'MASTER' and name contains 'ROB'") % parent_clause
    result = service.files().list(
        q=q, fields="files(id,name,parents)", pageSize=50,
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute()
    for f in result.get("files", []):
        return f["id"], f["name"]
    return None, "No ROB master file found in Drive."


def _is_rob_month_blank(ws, block_start):
    """Return True if cols 2,3,4 of the Revenue row are all empty."""
    rev_row = block_start + 1
    return all(
        ws.cell(rev_row, c).value is None or ws.cell(rev_row, c).value == ""
        for c in [2, 3, 4]
    )


def _rob_as_of_date(year, month):
    """Return the 1st of the month for a given year, advanced past Sat/Sun to Monday."""
    d = datetime.date(year, month, 1)
    if d.weekday() == 5:    # Saturday → Monday
        d += datetime.timedelta(days=2)
    elif d.weekday() == 6:  # Sunday → Monday
        d += datetime.timedelta(days=1)
    return d


def _resolve_cell(prev_wb_data, prev_wb_formulas, sheet_name, row, col):
    """Read a cell value, following simple cross-sheet formula references if needed.
    When data_only=True returns None (no cached value), reads the formula string
    from prev_wb_formulas and follows the reference (e.g. ='wk one'!B45)."""
    import re
    from openpyxl.utils import column_index_from_string

    def _read_data(wb, sname, r, c):
        if wb and sname in wb.sheetnames:
            return wb[sname].cell(r, c).value
        return None

    v = _read_data(prev_wb_data, sheet_name, row, col)
    if v is not None:
        return v

    # No cached value — try to follow the formula reference
    if prev_wb_formulas and sheet_name in prev_wb_formulas.sheetnames:
        formula = prev_wb_formulas[sheet_name].cell(row, col).value
        if formula and str(formula).startswith("="):
            m = re.match(r"^='?([^'!]+)'?!([A-Za-z]+)(\d+)$", str(formula).strip())
            if m:
                ref_sheet = m.group(1)
                ref_col   = column_index_from_string(m.group(2))
                ref_row   = int(m.group(3))
                v = _read_data(prev_wb_data, ref_sheet, ref_row, ref_col)
                if v is not None:
                    return v
                # Referenced sheet might itself have a formula — follow one more level
                v2 = _resolve_cell(prev_wb_data, prev_wb_formulas, ref_sheet, ref_row, ref_col)
                return v2
    return None


def _fill_rob_prev_table(wk1_ws, prev_wb, prev_wb_formulas, target_month):
    """Fill the 'Week 1 Previous Sheet - CALCULATION ONLY' table in wk1.
    Scans dynamically for header, year columns, and month rows.
    Pulls Revenue from last completed week tab of prev ROB."""

    month_abbrs = ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]

    def _as_year(v):
        """Try to extract a 4-digit year from a cell value (int, float, str, date)."""
        if isinstance(v, (datetime.datetime, datetime.date)):
            return v.year
        if isinstance(v, (int, float)):
            iv = int(v)
            if 2000 <= iv <= 2100:
                return iv
            # Excel date serial — convert
            if 40000 <= iv <= 60000:
                try:
                    d = datetime.date(1899, 12, 30) + datetime.timedelta(days=iv)
                    return d.year
                except Exception:
                    pass
        if isinstance(v, str):
            s = v.strip()
            if s.isdigit() and 2000 <= int(s) <= 2100:
                return int(s)
        return None

    # ── 1. Find header cell ───────────────────────────────────────────────────
    hdr_row = hdr_col = None
    for r in range(1, min(wk1_ws.max_row + 1, 60)):
        for c in range(1, wk1_ws.max_column + 1):
            val = str(wk1_ws.cell(r, c).value or "").strip().lower()
            if "week 1 previous" in val or ("calculation only" in val and "week" in val):
                hdr_row, hdr_col = r, c
                break
        if hdr_row:
            break
    if not hdr_row:
        return "Week 1 Previous Sheet table not found in wk one"

    years_in_order = [target_month.year - 3, target_month.year - 2,
                      target_month.year - 1, target_month.year]

    # ── 2. Find month label column first (Jan/Feb… to the LEFT of year cols) ─
    month_label_col = None
    for r in range(hdr_row + 1, hdr_row + 35):
        for c in range(max(1, hdr_col - 5), hdr_col + 5):
            v = str(wk1_ws.cell(r, c).value or "").strip().lower()
            if v in month_abbrs:
                month_label_col = c
                break
        if month_label_col:
            break
    if not month_label_col:
        return "Could not find month label column in Week 1 Previous Sheet table"

    # ── 3. Collect all month rows using that exact column ─────────────────────
    dest_month_row = {}  # month_idx (0-based) → row
    for r in range(hdr_row + 1, hdr_row + 35):
        v = str(wk1_ws.cell(r, month_label_col).value or "").strip().lower()
        if v in month_abbrs:
            dest_month_row[month_abbrs.index(v)] = r
    if not dest_month_row:
        return "Could not find month rows in Week 1 Previous Sheet table"

    # ── 4. Find year columns — scan to the RIGHT of month label col ───────────
    year_start_col = month_label_col + 1
    dest_year_col = {}  # year (int) → col in new wk1

    for dr in range(1, 8):
        # Try parsing actual year values
        parsed = {}
        for c in range(year_start_col, year_start_col + 20):
            yr = _as_year(wk1_ws.cell(hdr_row + dr, c).value)
            if yr:
                parsed[yr] = c
        if parsed:
            dest_year_col = parsed
            break
        # Fallback: find non-empty cells positionally (formulas count as non-empty)
        non_empty = []
        for c in range(year_start_col, year_start_col + 20):
            v = wk1_ws.cell(hdr_row + dr, c).value
            if v is not None and str(v).strip():
                non_empty.append(c)
        if len(non_empty) >= 2:
            for i, c in enumerate(non_empty[:4]):
                dest_year_col[years_in_order[i]] = c
            break

    if not dest_year_col:
        return "Could not find year columns in Week 1 Previous Sheet table"

    # ── 4. Build ordered list of week sheet names (most recent first) ────────
    wk_order = ["wk six", "wk five", "wk four", "wk three", "wk two", "wk one"]
    wk_sheet_names = []  # sheet names in prev ROB, most-recent first
    for wk_try in wk_order:
        matches = [s for s in prev_wb.sheetnames if wk_try in s.lower()]
        if matches:
            wk_sheet_names.append(matches[0])
    if not wk_sheet_names:
        return "No week tabs found in previous ROB"

    # ── 5. Year → source col: scan wk one row 4 for dates/years ─────────────
    base_year = target_month.year
    src_year_col = {base_year - 3: 2, base_year - 2: 3, base_year - 1: 4, base_year: 5}

    wk1_sheet_name = wk_sheet_names[-1]  # wk one is last in most-recent-first list
    ref_ws = prev_wb[wk1_sheet_name]
    detected = {}
    for c in range(1, min(ref_ws.max_column + 1, 20)):
        yr = _as_year(ref_ws.cell(4, c).value)
        if yr and 2000 <= yr <= 2100:
            detected[yr] = c
    if len(detected) >= 3:
        src_year_col = detected

    # ── 6. Write Revenue values — check each cell individually ───────────────
    # For each (month, year) cell: walk week tabs most-recent→oldest.
    # Each cell is checked on its own — some are hardcoded numbers, some are
    # formulas pointing elsewhere. _resolve_cell handles both cases.
    for month_idx, dest_row in dest_month_row.items():
        rev_row = 4 + 8 * month_idx + 1
        for year, dest_col in dest_year_col.items():
            src_col = src_year_col.get(year)
            if not src_col:
                continue
            v = None
            for sheet_name in wk_sheet_names:
                v = _resolve_cell(prev_wb, prev_wb_formulas, sheet_name, rev_row, src_col)
                if v is not None and not is_formula(str(v)):
                    break  # found a real value in this tab — use it
                v = None   # reset if None or formula string slipped through
            if v is not None:
                wk1_ws.cell(dest_row, dest_col).value = v

    return None


def _fill_rob_sheet(new_ws, prev_ws, ly_ws, target_month, is_wk_one, wk_one_sheet_name):
    """Fill one ROB sheet tab with historical data."""
    from openpyxl.utils import get_column_letter

    target_idx  = target_month.month - 1   # 0-based (Jul = 6)
    prev_idx    = target_idx - 1           # most recently completed month (Jun = 5)
    # LY col → new col shift: LY has [2022,2023,2024,2025], new needs [2023,2024,2025,2026]
    ly_to_new    = {3: 2, 4: 3, 5: 4}
    data_offsets = [1, 2, 3, 4, 5, 6, 7]  # offset 0 (date header) handled separately

    # ── Build as-of dates once from LY ROB (same for every month block) ─────
    # LY ROB cols 3,4,5 hold the prior 3 years' reporting dates → new cols 2,3,4
    # Col 5 (current year) = the 1st of the target month
    as_of_dates = {}
    if ly_ws:
        ref_row = 4 + 8 * target_idx   # any block works; use target month block
        for ly_col, new_col in ly_to_new.items():
            v = ly_ws.cell(ref_row, ly_col).value
            if v is not None and not is_formula(str(v)):
                as_of_dates[new_col] = v
    as_of_dates[5] = datetime.datetime(target_month.year, target_month.month, target_month.day)

    for month_idx in range(12):
        block_start = 4 + 8 * month_idx

        # ── Date header row (offset 0): always write from computed as_of_dates ──
        if is_wk_one:
            for col, date_val in as_of_dates.items():
                new_ws.cell(block_start, col).value = date_val
        else:
            for col in as_of_dates:
                col_ltr = get_column_letter(col)
                new_ws.cell(block_start, col).value = f"='{wk_one_sheet_name}'!{col_ltr}{block_start}"

        # ── Data rows (offsets 1–7) ───────────────────────────────────────────
        if month_idx < prev_idx:
            # Past months (Jan–May when target=Jul): copy all 4 cols from prev ROB
            if not _is_rob_month_blank(new_ws, block_start):
                continue
            if is_wk_one:
                if prev_ws is None:
                    continue
                for dr in data_offsets:
                    r = block_start + dr
                    for c in [2, 3, 4, 5]:
                        v = prev_ws.cell(r, c).value
                        if v is not None and not is_formula(str(v)) and not is_datelike(v):
                            new_ws.cell(r, c).value = v
            else:
                for dr in data_offsets:
                    r = block_start + dr
                    for c in [2, 3, 4, 5]:
                        col_ltr = get_column_letter(c)
                        new_ws.cell(r, c).value = f"='{wk_one_sheet_name}'!{col_ltr}{r}"

        elif month_idx == prev_idx:
            # Prev month (Jun when target=Jul):
            # Cols 2,3,4 = historical years from LY ROB (same source as Jul+)
            # Col 5     = actual current-year data from prev ROB (built up weekly)
            if not _is_rob_month_blank(new_ws, block_start):
                continue
            if is_wk_one:
                if ly_ws:
                    for dr in data_offsets:
                        r = block_start + dr
                        for ly_col, new_col in ly_to_new.items():
                            v = ly_ws.cell(r, ly_col).value
                            if v is not None and not is_formula(str(v)) and not is_datelike(v):
                                new_ws.cell(r, new_col).value = v
                    ly_sec_col = find_secondary_col(ly_ws, block_start) or 7
                    for dr in [4, 5, 6]:
                        r = block_start + dr
                        v = ly_ws.cell(r, ly_sec_col).value
                        if v is not None and not is_formula(str(v)) and not is_datelike(v):
                            new_ws.cell(r, 8).value = v
            else:
                for dr in data_offsets:
                    r = block_start + dr
                    for c in [2, 3, 4, 5]:
                        col_ltr = get_column_letter(c)
                        new_ws.cell(r, c).value = f"='{wk_one_sheet_name}'!{col_ltr}{r}"

        else:
            # Current month and future months (Jul+): cols 2,3,4 from LY ROB
            if ly_ws is None:
                continue
            for dr in data_offsets:
                r = block_start + dr
                for ly_col, new_col in ly_to_new.items():
                    v = ly_ws.cell(r, ly_col).value
                    if v is not None and not is_formula(str(v)) and not is_datelike(v):
                        new_ws.cell(r, new_col).value = v
            ly_sec_col = find_secondary_col(ly_ws, block_start) or 7
            for dr in [4, 5, 6]:
                r = block_start + dr
                v = ly_ws.cell(r, ly_sec_col).value
                if v is not None and not is_formula(str(v)) and not is_datelike(v):
                    new_ws.cell(r, 8).value = v


def setup_new_rob_month(service, hotel_id: str, hotel_name: str, target_month: datetime.date):
    """Full ROB new-month setup. Returns (new_file_name, error_str)."""
    year_kw  = str(target_month.year)
    month_kw = target_month.strftime("%b%Y").upper()

    # ── Locate or create month folder ────────────────────────────────────────
    rev_id, _ = _find_rev_reports_folder_for_year(service, hotel_id, year_kw)
    if not rev_id:
        return None, "No REVENUE REPORTS folder."
    month_id, _ = _find_or_create_month_folder_under_rev(service, rev_id, year_kw, month_kw, target_month, hotel_name)

    # ── Find or copy the file ─────────────────────────────────────────────────
    existing_id, existing_name = drive_find_file(service, "ROB", month_id)
    is_fresh_copy = False
    if existing_id and "master" not in existing_name.lower():
        new_file_id, new_file_name = existing_id, existing_name
    else:
        is_fresh_copy = True
        master_id, master_name = find_rob_master(service, hotel_id)
        if not master_id:
            return None, master_name
        hotel_suffix = hotel_name.upper()
        name_upper   = master_name.upper()
        if "ROB" in name_upper:
            after = master_name[name_upper.find("ROB") + 3:].strip()
            after = after.replace(".xlsx","").replace(".xlsm","").replace(".XLSX","").replace(".XLSM","").strip()
            if after:
                hotel_suffix = after
        ext = ".xlsm" if master_name.lower().endswith(".xlsm") else ".xlsx"
        new_file_name = f"{month_kw} ROB {hotel_suffix}{ext}"
        try:
            new_file_id, new_file_name = drive_copy_file(service, master_id, new_file_name, month_id)
        except Exception as e:
            return None, str(e)

    # ── Load all three workbooks ──────────────────────────────────────────────
    new_wb_bytes = drive_download(service, new_file_id)
    new_wb = openpyxl.load_workbook(io.BytesIO(new_wb_bytes), data_only=False)
    if is_fresh_copy:
        clear_tab_colors(new_wb, ROB_SHEETS)

    # Resolution/load failures here used to be swallowed silently — the
    # ROB workbook would just come back with July onward blank and no
    # indication why. Both lookup failures (resolve_drive_workbook returning
    # an error) and load failures (download/openpyxl exceptions) are now
    # captured into `warnings` and surfaced to the caller alongside the
    # success message, matching the diagnostic the SR flow already shows.
    warnings = []

    prev_month_dt = (target_month - datetime.timedelta(days=1)).replace(day=1)
    prev_result, prev_err = resolve_drive_workbook(service, hotel_id, hotel_name, "ROB",
                                                     month_date=prev_month_dt)
    prev_wb = None
    prev_wb_formulas = None
    if prev_result:
        try:
            prev_bytes = drive_download(service, prev_result[0])
            prev_wb = openpyxl.load_workbook(io.BytesIO(prev_bytes), data_only=True)
            prev_wb_formulas = openpyxl.load_workbook(io.BytesIO(prev_bytes), data_only=False)
        except Exception as e:
            warnings.append(f"Prev month ({prev_month_dt.strftime('%b %Y')}) workbook found but failed to load: {e}")
    else:
        warnings.append(f"Prev month ({prev_month_dt.strftime('%b %Y')}) not found: {prev_err}")

    ly_month_dt = target_month.replace(year=target_month.year - 1)
    ly_result, ly_err = resolve_drive_workbook(service, hotel_id, hotel_name, "ROB",
                                                month_date=ly_month_dt)
    ly_wb = None
    if ly_result:
        try:
            ly_wb = openpyxl.load_workbook(
                io.BytesIO(drive_download(service, ly_result[0])), data_only=True)
        except Exception as e:
            warnings.append(f"Last year ({ly_month_dt.strftime('%b %Y')}) workbook found but failed to load: {e}")
    else:
        warnings.append(f"Last year ({ly_month_dt.strftime('%b %Y')}) not found — future months' historical "
                         f"columns will be blank: {ly_err}")

    # ── Fill each sheet ───────────────────────────────────────────────────────
    wk_one_name = ROB_SHEETS[0]
    for sheet_name in ROB_SHEETS:
        if sheet_name not in new_wb.sheetnames:
            continue
        new_ws  = new_wb[sheet_name]
        prev_ws = prev_wb[sheet_name] if prev_wb and sheet_name in prev_wb.sheetnames else None
        ly_ws   = ly_wb[sheet_name]   if ly_wb   and sheet_name in ly_wb.sheetnames   else None
        is_wk_one = (sheet_name == wk_one_name)
        _fill_rob_sheet(new_ws, prev_ws, ly_ws, target_month, is_wk_one, wk_one_name)

    # ── Fill Week 1 Previous Sheet table in wk one ───────────────────────────
    if prev_wb and wk_one_name in new_wb.sheetnames:
        err = _fill_rob_prev_table(new_wb[wk_one_name], prev_wb, prev_wb_formulas, target_month)
        if err:
            warnings.append(f"Prev table: {err}")

    strip_tables(new_wb)
    out = io.BytesIO()
    new_wb.save(out)
    drive_upload(service, new_file_id, out.getvalue(), new_file_name)
    warn_str = "; ".join(warnings) if warnings else None
    return new_file_name, warn_str


def find_forecast_master(service, hotel_id: str):
    """Search the hotel's own Drive tree for the Forecast master file (.xlsx or .xlsm)."""
    scope_ids = _hotel_search_scope_ids(service, hotel_id)
    if not scope_ids:
        return None, "Could not resolve hotel folder to search."
    parent_clause = " or ".join("'%s' in parents" % sid for sid in scope_ids)
    q = ("trashed=false and (%s) "
         "and (mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' "
         "or mimeType='application/vnd.ms-excel.sheet.macroenabled.12') "
         "and name contains 'MASTER' and name contains 'FORECAST'") % parent_clause
    result = service.files().list(
        q=q, fields="files(id,name,parents)", pageSize=50,
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute()
    for f in result.get("files", []):
        return f["id"], f["name"]
    return None, "No FORECAST master file found in Drive."


def setup_new_forecast_month(service, hotel_id: str, hotel_name: str, target_month: datetime.date):
    """
    Copy Forecast master → rename for target_month → place in month folder → set B4 date.
    Returns (new_file_name, error_str).
    """
    year_kw  = str(target_month.year)
    month_kw = target_month.strftime("%b%Y").upper()

    rev_id, _ = _find_rev_reports_folder_for_year(service, hotel_id, year_kw)
    if not rev_id:
        return None, "No REVENUE REPORTS folder."

    month_id, _ = _find_or_create_month_folder_under_rev(service, rev_id, year_kw, month_kw, target_month, hotel_name)

    # Check if Forecast already exists
    existing_id, existing_name = drive_find_file(service, "FORECAST", month_id)
    if existing_id and "master" not in existing_name.lower():
        return existing_name, None

    master_id, master_name = find_forecast_master(service, hotel_id)
    if not master_id:
        return None, master_name

    # Infer hotel suffix from master name
    hotel_suffix = hotel_name.upper()
    name_upper = master_name.upper()
    ext = ".xlsm" if master_name.lower().endswith(".xlsm") else ".xlsx"
    for kw in ("FORECAST",):
        if kw in name_upper:
            after = master_name[name_upper.find(kw) + len(kw):].strip()
            after = after.replace(".xlsx","").replace(".xlsm","").replace(".XLSX","").replace(".XLSM","").strip()
            if after:
                hotel_suffix = after
            break

    new_file_name = f"{month_kw} FORECAST {hotel_suffix}{ext}"
    try:
        new_file_id, created_name = drive_copy_file(service, master_id, new_file_name, month_id)
    except Exception as e:
        return None, str(e)

    # Set B4 = first day of target_month in FCST-WK1
    try:
        wb_bytes  = drive_download(service, new_file_id)
        wb        = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=False)
        clear_tab_colors(wb, FORECAST_SHEETS)
        sheet     = FORECAST_SHEETS[0] if FORECAST_SHEETS[0] in wb.sheetnames else wb.sheetnames[1] if len(wb.sheetnames) > 1 else wb.sheetnames[0]
        ws        = wb[sheet]
        # Find "Day of Week" cell → one right + one down = start date cell
        date_cell_row = date_cell_col = None
        for r in range(1, 15):
            for c in range(1, 10):
                if "day of week" in str(ws.cell(r, c).value or "").lower():
                    date_cell_row = r + 1
                    date_cell_col = c + 1
                    break
            if date_cell_row:
                break
        if date_cell_row and date_cell_col:
            ws.cell(date_cell_row, date_cell_col).value = datetime.datetime(
                target_month.year, target_month.month, 1)
        strip_tables(wb)
        out = io.BytesIO()
        wb.save(out)
        drive_upload(service, new_file_id, out.getvalue(), created_name)
    except Exception as e:
        return created_name, f"Copied OK but could not set start date: {e}"

    return created_name, None


def find_sr_master(service, hotel_id: str):
    """Search the hotel's own Drive tree for the SR master file.
    Returns (file_id, file_name) or (None, error_str).
    """
    scope_ids = _hotel_search_scope_ids(service, hotel_id)
    if not scope_ids:
        return None, "Could not resolve hotel folder to search."
    parent_clause = " or ".join("'%s' in parents" % sid for sid in scope_ids)
    q = ("trashed=false and (%s) "
         "and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' "
         "and name contains 'MASTER' and name contains 'STRATEGY'") % parent_clause
    result = service.files().list(
        q=q, fields="files(id,name,parents)", pageSize=50,
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute()
    for f in result.get("files", []):
        return f["id"], f["name"]
    return None, "No STRATEGY master file found in Drive."


def setup_new_sr_month(service, hotel_id: str, hotel_name: str, target_month: datetime.date):
    """
    New month SR setup:
      1. Find SR master, copy it, rename to [MON][YEAR] STRATEGY [HOTEL].xlsx
      2. Find or create month folder under the year folder
      3. Move the copy into that folder
    Returns (new_file_name, error_str).
    """
    year_kw = str(target_month.year)
    month_kw = target_month.strftime("%b%Y").upper()

    # Resolve REVENUE REPORTS and month folder
    rev_id, _ = _find_rev_reports_folder_for_year(service, hotel_id, year_kw)
    if not rev_id:
        return None, "No REVENUE REPORTS folder."

    # Find or create month folder
    month_id, month_name = _find_or_create_month_folder_under_rev(service, rev_id, year_kw, month_kw, target_month, hotel_name)

    # Check if SR already exists in that folder
    existing_id, existing_name = drive_find_file(service, "STRATEGY", month_id)
    if existing_id and "master" not in existing_name.lower():
        return existing_name, None  # already set up

    # Find master
    master_id, master_name = find_sr_master(service, hotel_id)
    if not master_id:
        return None, master_name  # error string

    # Infer hotel suffix from master file name for the new file name
    # e.g. "MASTER 2026 STRATEGY PLYMOUTH.xlsx" → "PLYMOUTH"
    hotel_suffix = hotel_name.upper()
    name_upper = master_name.upper().replace(".XLSX", "")
    if "STRATEGY" in name_upper:
        after = name_upper[name_upper.find("STRATEGY") + len("STRATEGY"):].strip()
        if after:
            hotel_suffix = master_name[master_name.upper().find("STRATEGY") + len("STRATEGY"):].strip().replace(".xlsx", "").replace(".XLSX", "").strip()

    new_file_name = f"{month_kw} STRATEGY {hotel_suffix}.xlsx"
    try:
        _, created_name = drive_copy_file(service, master_id, new_file_name, month_id)
    except Exception as e:
        return None, str(e)
    return created_name, None


def get_prev_month_otb_trans(service, hotel_id: str, hotel_name: str, current_month: datetime.date):
    """Pull OTB TY Trans values for current_month dates from previous month's SR last filled tab.
    Returns {date: value} or {} on any failure.
    """
    prev_month = (current_month.replace(day=1) - datetime.timedelta(days=1)).replace(day=1)
    result, err = resolve_drive_workbook(service, hotel_id, hotel_name, "Strategy Report", month_date=prev_month)
    if err or not result:
        return {}

    file_id, _ = result
    try:
        file_bytes = drive_download(service, file_id)
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    except Exception:
        return {}

    # Find last filled (colored) tab
    last_filled_sheet = None
    for sheet_name in STRATEGY_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        tab_color = ws.sheet_properties.tabColor
        if tab_color is not None:
            last_filled_sheet = sheet_name

    if not last_filled_sheet:
        return {}

    ws = wb[last_filled_sheet]
    col_map = detect_strategy_columns(ws)
    otb_col = col_map.get("otb_trans")
    date_col = detect_date_column(ws)
    if not otb_col:
        return {}

    result_map = {}
    for r in range(5, ws.max_row + 1):
        date_val = ws.cell(r, date_col).value
        if isinstance(date_val, datetime.datetime):
            d = date_val.date()
        elif isinstance(date_val, datetime.date):
            d = date_val
        else:
            continue
        if d < current_month:
            continue
        cell_val = ws.cell(r, otb_col).value
        if cell_val is not None and not is_formula(cell_val):
            result_map[d] = safe_float(cell_val)

    return result_map


def resolve_drive_workbook(service, hotel_id: str, hotel_name: str, workbook_type: str, month_date: datetime.date = None):
    """
    Walk Drive to find the target workbook. Handles two folder structures:
      A) Hotel > MMMYYYY REVENUE REPORTS HOTEL > files  (month in folder name, files direct)
      B) Hotel > REVENUE REPORTS > Year > Month > files (nested year/month subfolders)
    Returns ((file_id, file_name), None) or (None, error_message).
    Never touches files whose name contains 'master'.

    A "hotel" can also be a MULTI:<id>,<id>,... group — several candidate
    root folders sharing one dropdown entry, either because they're each a
    flat per-year/month folder shared directly (Hyannis Anchor In) or full
    duplicate top-level hotel folders from historical typos/copies
    (confirmed real case: "Provinceetown Surfside", "Provincertown
    Surfside", "Surfside (1)" all being the same hotel). Each candidate is
    resolved fully (structures A/B/C, recursing into its own REVENUE REPORTS
    child if it has one) rather than guessed from its name alone — whichever
    candidate actually contains the target file wins.
    """
    if month_date is None:
        month_date = datetime.date.today()

    month_kw        = month_date.strftime("%b%Y").upper()
    month_kw_2digit = month_date.strftime("%b%y").upper()
    year_kw         = str(month_date.year)
    wb_keyword      = WORKBOOK_KEYWORDS[workbook_type]

    def _list_subfolders(parent_id):
        q = (f"'{parent_id}' in parents and trashed = false and "
             f"mimeType = 'application/vnd.google-apps.folder'")
        return service.files().list(
            q=q, fields="files(id, name)", pageSize=100,
            supportsAllDrives=True, includeItemsFromAllDrives=True,
        ).execute().get("files", [])

    def _find_file_in(folder_id, folder_name):
        fid, fname = drive_find_file(service, wb_keyword, folder_id)
        if not fid:
            return None, f"No '{wb_keyword}' workbook found in '{folder_name}'."
        if "master" in fname.lower():
            return None, f"Resolved file '{fname}' looks like a master doc — aborting."
        return (fid, fname), None

    def _resolve_single(single_id, single_name):
        """Resolve within ONE candidate root — either a full hotel-parent
        folder (with its own REVENUE REPORTS > year > month nesting) or a
        folder that IS already the REVENUE REPORTS level directly (detected
        from its own name). Returns ((file_id, file_name), None) or (None, err).
        """
        self_is_rev = _is_rev_reports_name(single_name)
        children = _list_subfolders(single_id)

        # A: Hotel > MMMYYYY REVENUE REPORTS HOTEL > file
        if not self_is_rev:
            a = next((f for f in children
                       if _is_rev_reports_name(f["name"]) and month_kw in f["name"].upper()), None)
            if a:
                result = _find_file_in(a["id"], a["name"])
                if result[0]:
                    return result

        # B: Hotel > REVENUE REPORTS > MMMYYYY ... > file
        if self_is_rev:
            rev = {"id": single_id, "name": single_name}
        else:
            rev = next((f for f in children
                        if _is_rev_reports_name(f["name"]) and year_kw in f["name"]), None)
            if not rev:
                rev = next((f for f in children if _is_rev_reports_name(f["name"])), None)
        if rev:
            rev_children = children if self_is_rev else _list_subfolders(rev["id"])
            b1 = next((f for f in rev_children if month_kw in f["name"].upper()), None)
            if b1:
                result = _find_file_in(b1["id"], b1["name"])
                if result[0]:
                    return result
            b2_year = next((f for f in rev_children if year_kw in f["name"].upper()), None)
            if b2_year:
                b2_month_id, b2_month_name = drive_find_folder_by_keyword(
                    service, month_kw, parent_id=b2_year["id"])
                if b2_month_id:
                    result = _find_file_in(b2_month_id, b2_month_name)
                    if result[0]:
                        return result

        # C: Hotel > Year > Month > file  (no REVENUE REPORTS wrapper)
        c_year = next((f for f in children
                       if year_kw in f["name"].upper()
                       and not _is_rev_reports_name(f["name"])), None)
        if c_year:
            c_month_id, c_month_name = drive_find_folder_by_keyword(
                service, month_kw, parent_id=c_year["id"])
            if c_month_id:
                result = _find_file_in(c_month_id, c_month_name)
                if result[0]:
                    return result

        return None, f"Could not find '{month_kw}' workbook under '{single_name}'."

    if hotel_id.startswith(MULTI_ID_PREFIX):
        candidate_ids = hotel_id[len(MULTI_ID_PREFIX):].split(",")
        candidates = []
        for cid in candidate_ids:
            try:
                info = service.files().get(fileId=cid, fields="name", supportsAllDrives=True).execute()
                candidates.append({"id": cid, "name": info["name"]})
            except Exception:
                continue
        if not candidates:
            return None, f"Could not read any of the shared folders for '{hotel_name}'."

        # Try the most likely-named candidates first (pure ordering
        # optimization), but fall through to the next candidate if one
        # doesn't actually contain the file — e.g. a stale duplicate folder
        # from a typo — instead of giving up after the first name match.
        def _sort_key(f):
            name_up = f["name"].upper()
            if month_kw in name_up or month_kw_2digit in name_up:
                return 0
            if year_kw in name_up:
                return 1
            return 2
        ordered = sorted(candidates, key=_sort_key)

        last_err = None
        for cand in ordered:
            result, err = _resolve_single(cand["id"], cand["name"])
            if result:
                return result, None
            last_err = err
        return None, last_err or f"Could not find '{month_kw}' workbook for '{hotel_name}'."

    return _resolve_single(hotel_id, hotel_name)


DOW_ABBREVS = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]


def _count_sheet_data_rows(ws):
    """Count how many data rows exist from row 5 downward by finding the last
    row that has any non-empty content across the first 10 columns."""
    last_row = 4
    for r in range(5, ws.max_row + 1):
        if any(ws.cell(r, c).value is not None for c in range(1, 11)):
            last_row = r
    return max(0, last_row - 4)  # number of rows starting from row 5


def restructure_sr_dates(wb, target_month):
    """Restructure the three date columns in every strategy sheet for a new month.
    Row 4 (header):    col 1 = LY year, col 3 = TY year — else stale years carry
                        over from whatever the master template last had.
    Col 1 (LY date):   starts at day 2 of target_month, last year
    Col 2 (day of wk): abbreviation matching the TY date in col 3
    Col 3 (TY date):   starts at day 1 of target_month, this year
    Row count matches the master — hotels open only part of the year have fewer rows.
    """
    ty_start = target_month
    ly_start = datetime.date(ty_start.year - 1, ty_start.month, 2)

    for sheet_name in STRATEGY_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        ws.cell(4, 1).value = ly_start.year
        ws.cell(4, 3).value = ty_start.year
        num_rows = min(_count_sheet_data_rows(ws), 365)
        if num_rows == 0:
            continue
        for i in range(num_rows):
            row     = 5 + i
            ty_date = ty_start + datetime.timedelta(days=i)
            ly_date = ly_start + datetime.timedelta(days=i)
            dow     = DOW_ABBREVS[ty_date.weekday()]
            ws.cell(row, 1).value = datetime.datetime(ly_date.year, ly_date.month, ly_date.day)
            ws.cell(row, 2).value = dow
            ws.cell(row, 3).value = datetime.datetime(ty_date.year, ty_date.month, ty_date.day)


def _load_wb_from_drive(svc, hotel_id, hotel_name, wb_type, month_date, data_only=True):
    """Download and parse a workbook from Drive. Returns openpyxl.Workbook or None.
    data_only=True (default) returns cached cell values — use for reference workbooks.
    data_only=False preserves formulas — use for workbooks we intend to write back.
    """
    result, err = resolve_drive_workbook(svc, hotel_id, hotel_name, wb_type, month_date=month_date)
    if err or not result:
        return None
    try:
        return openpyxl.load_workbook(io.BytesIO(drive_download(svc, result[0])), data_only=data_only)
    except Exception:
        return None


# ── User accounts (self-serve requests + admin approval) ─────────────────────
# Streamlit Cloud's filesystem is ephemeral and st.secrets is read-only at
# runtime, so per-person accounts created through the app can't live in either
# — store them in a small JSON file in Drive instead, using the exact same
# download/upload plumbing already used for every workbook in this app.

APP_DATA_FOLDER_NAME = "Workbook Updater App Data"
USERS_FILE_NAME      = "users.json"


def _find_app_data_folder(service):
    """Find the shared App Data folder — matched case-insensitively against
    APP_DATA_FOLDER_NAME (confirmed real case: exact-match search missed a
    folder created as 'workbook updater app data', different casing than the
    original name). Must be created once and shared with the service
    account, the same way every hotel's Drive folder is — can live anywhere
    inside a Shared Drive, including nested inside a hotel's folder."""
    q = "mimeType = 'application/vnd.google-apps.folder' and trashed = false"
    result = service.files().list(
        q=q, fields="files(id, name)", pageSize=1000,
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute()
    target = APP_DATA_FOLDER_NAME.strip().lower()
    for f in result.get("files", []):
        if f["name"].strip().lower() == target:
            return f["id"]
    return None


def _find_or_create_users_file(service):
    """Return (file_id, error). error is a user-facing string when the App
    Data folder itself hasn't been created/shared yet — creates users.json
    inside it (empty) on first use otherwise."""
    folder_id = _find_app_data_folder(service)
    if not folder_id:
        return None, (f"Drive folder '{APP_DATA_FOLDER_NAME}' not found. Create it in "
                       f"Google Drive and share it with the service account (same as a "
                       f"hotel folder), then try again.")
    q = "'%s' in parents and trashed = false and name = '%s'" % (folder_id, USERS_FILE_NAME)
    result = service.files().list(
        q=q, fields="files(id, name)", pageSize=5,
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute()
    files = result.get("files", [])
    if files:
        return files[0]["id"], None
    empty = json.dumps({"users": []}).encode("utf-8")
    media = MediaIoBaseUpload(io.BytesIO(empty), mimetype="application/json", resumable=True)
    created = service.files().create(
        body={"name": USERS_FILE_NAME, "parents": [folder_id]},
        media_body=media, fields="id", supportsAllDrives=True,
    ).execute()
    return created["id"], None


def _load_users(service, file_id):
    """Return the list of user dicts from users.json (empty list if unreadable)."""
    try:
        raw = drive_download(service, file_id)
        return json.loads(raw.decode("utf-8")).get("users", [])
    except Exception:
        return []


def _save_users(service, file_id, users):
    payload = json.dumps({"users": users}, indent=2).encode("utf-8")
    drive_upload(service, file_id, payload, USERS_FILE_NAME)


# ── UI ────────────────────────────────────────────────────────────────────────

st.set_page_config(page_title="Linchris Weekly Tools", layout="wide")


# ── Login gate ────────────────────────────────────────────────────────────────
def check_login(username: str, password: str):
    """Return (ok, is_admin). Checks the single admin account (Streamlit
    secrets) first, then falls back to approved entries in users.json."""
    admin_user = st.secrets["auth"]["username"]
    admin_hash = st.secrets["auth"]["password_hash"].encode()
    if username == admin_user and bcrypt.checkpw(password.encode(), admin_hash):
        return True, True

    try:
        svc = get_drive_service()
        file_id, err = _find_or_create_users_file(svc)
        if err:
            return False, False
        users = _load_users(svc, file_id)
    except Exception:
        return False, False

    for u in users:
        if (u.get("username") == username and u.get("status") == "approved"
                and bcrypt.checkpw(password.encode(), u.get("password_hash", "").encode())):
            return True, u.get("role") == "admin"
    return False, False


LOGIN_ENABLED = True
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False
    st.session_state["is_admin"] = False
    st.session_state["username"] = None

if LOGIN_ENABLED and not st.session_state["authenticated"]:
    st.title("Linchris Hotel Corporation")
    st.subheader("Please log in to continue")

    login_tab, access_key_tab = st.tabs(["Log In", "Access Key"])

    with login_tab:
        with st.form("login_form"):
            username = st.text_input("Username")
            password = st.text_input("Password", type="password")
            submitted = st.form_submit_button("Log In")
            if submitted:
                ok, is_admin = check_login(username, password)
                if ok:
                    st.session_state["authenticated"] = True
                    st.session_state["is_admin"] = is_admin
                    st.session_state["username"] = username
                    st.rerun()
                else:
                    st.error("Incorrect username or password — or your account is still pending admin approval.")

    with access_key_tab:
        # Temporary, simple stopgap: a single shared secret (Streamlit
        # secrets, no Drive/storage involved) grants standard access —
        # no per-person account, no folder setup. Fine for onboarding a
        # handful of people quickly; can be replaced with a real per-person
        # account system (already built once, just dormant — see
        # _find_or_create_users_file / render_admin_settings) later.
        st.caption("Enter the access key given to you by an admin to get in.")
        with st.form("access_key_form"):
            access_key_input = st.text_input("Access key", type="password", key="access_key_input")
            display_name     = st.text_input("Your name", key="access_key_display_name")
            key_submitted     = st.form_submit_button("Enter")
            if key_submitted:
                configured_key = st.secrets.get("auth", {}).get("access_key")
                if not configured_key:
                    st.error("Access key login isn't set up yet — an admin needs to add "
                             "'access_key' under [auth] in this app's secrets.")
                elif not access_key_input:
                    st.error("Access key is required.")
                elif access_key_input.strip() != configured_key:
                    st.error("Incorrect access key.")
                else:
                    st.session_state["authenticated"] = True
                    st.session_state["is_admin"] = False
                    st.session_state["username"] = display_name.strip() or "Guest"
                    st.rerun()

    st.stop()

st.markdown("""
<style>
  .block-container { max-width: 100% !important; padding-left: 2rem !important; padding-right: 2rem !important; }

  .stTabs [data-baseweb="tab-list"] { gap: 8px; border-bottom: 1px solid #E5E7EB; }
  .stTabs [data-baseweb="tab"] {
    background: #F1F3F5; color: #1E293B;
    border-radius: 6px 6px 0 0; padding: 8px 20px; font-weight: 600;
  }
  .stTabs [aria-selected="true"] {
    background: #2563EB !important; color: #FFFFFF !important;
    box-shadow: inset 0 -3px 0 #C9A84C;
  }
  div[data-testid="metric-container"] {
    background: #F8F9FA; border: 1px solid #E5E7EB; border-left: 3px solid #C9A84C;
    border-radius: 8px; padding: 12px;
  }
  button[data-testid="stBaseButton-pills"],
  button[data-testid="stBaseButton-pillsActive"] {
    border-radius: 8px !important;
    border: 1.5px solid #94A3B8 !important;
    font-size: 1.5rem !important;
    padding: 1rem 1.8rem !important;
  }
  button[data-testid="stBaseButton-pillsActive"] {
    background-color: #2563EB !important;
    border-color: #2563EB !important;
    color: #FFFFFF !important;
  }

  /* Larger, consistent widget labels + selectbox text app-wide */
  label[data-testid="stWidgetLabel"] p { font-size: 1.2rem !important; }
  div[data-testid="stSelectbox"] div[data-baseweb="select"] div { font-size: 1.25rem !important; }
  [data-testid="stFileUploaderDropzone"] { font-size: 1.15rem !important; }
  [data-testid="stFileUploaderDropzoneInstructions"] span { font-size: 1.15rem !important; }
  ul[data-testid="stSelectboxVirtualDropdown"] li {
    font-size: 1.25rem !important;
    -webkit-font-smoothing: antialiased;
    text-rendering: optimizeLegibility;
  }
</style>
""", unsafe_allow_html=True)

if "view" not in st.session_state:
    st.session_state["view"] = "main"


def render_admin_settings(svc, users_file_id, users_err):
    st.title("Admin Settings")
    if st.button("← Back to app"):
        st.session_state["view"] = "main"
        st.rerun()

    if users_err:
        st.warning(f"Account requests unavailable: {users_err}")
        return

    all_users  = _load_users(svc, users_file_id)
    admin_user = st.secrets["auth"]["username"]

    st.subheader("Pending Requests")
    pending = [u for u in all_users if u.get("status") == "pending"]
    if not pending:
        st.caption("No pending requests.")
    for u in pending:
        c1, c2, c3, c4 = st.columns([3, 3, 1, 1])
        c1.write(f"**{u.get('username')}**")
        c2.write(u.get("display_name") or "—")
        if c3.button("Approve", key=f"approve_{u.get('username')}"):
            for uu in all_users:
                if uu.get("username") == u.get("username"):
                    uu["status"] = "approved"
                    uu["decided_at"] = datetime.datetime.now().isoformat()
            _save_users(svc, users_file_id, all_users)
            st.rerun()
        if c4.button("Reject", key=f"reject_{u.get('username')}"):
            all_users = [uu for uu in all_users if uu.get("username") != u.get("username")]
            _save_users(svc, users_file_id, all_users)
            st.rerun()

    st.divider()

    st.subheader("All Users")
    rows = [{"Username": admin_user, "Name": "—", "Role": "Admin", "Status": "approved"}]
    for u in all_users:
        rows.append({
            "Username": u.get("username"),
            "Name":     u.get("display_name") or "—",
            "Role":     "Admin" if u.get("role") == "admin" else "Editor",
            "Status":   u.get("status"),
        })
    st.dataframe(rows, use_container_width=True)

    approved = [u for u in all_users if u.get("status") == "approved"]
    if approved:
        st.caption("Change role / revoke access:")
        for u in approved:
            rc1, rc2, rc3 = st.columns([4, 2, 1])
            rc1.write(f"{u.get('username')} ({u.get('display_name') or '—'})")
            current_role = "Admin" if u.get("role") == "admin" else "Editor"
            new_role = rc2.selectbox("Role", ["Editor", "Admin"],
                                      index=["Editor", "Admin"].index(current_role),
                                      key=f"role_{u.get('username')}", label_visibility="collapsed")
            if new_role.lower() != u.get("role", "editor"):
                for uu in all_users:
                    if uu.get("username") == u.get("username"):
                        uu["role"] = new_role.lower()
                _save_users(svc, users_file_id, all_users)
                st.rerun()
            if rc3.button("Revoke", key=f"revoke_{u.get('username')}"):
                all_users = [uu for uu in all_users if uu.get("username") != u.get("username")]
                _save_users(svc, users_file_id, all_users)
                st.rerun()

    st.divider()

    st.subheader("Add New User")
    st.caption("Creates and approves an account directly — skips the request/approval step.")
    with st.form("admin_add_user_form"):
        new_username = st.text_input("Username", key="admin_new_username")
        new_display  = st.text_input("Name", key="admin_new_display")
        new_password = st.text_input("Password", type="password", key="admin_new_password")
        new_role     = st.selectbox("Role", ["Editor", "Admin"], key="admin_new_role",
                                     help="Editor: normal app access, no Admin Settings. Admin: full access including this page.")
        submitted = st.form_submit_button("Add User")
        if submitted:
            if not new_username or not new_password:
                st.error("Username and password are required.")
            elif new_username == admin_user or any(u.get("username") == new_username for u in all_users):
                st.error("That username is already taken.")
            else:
                all_users.append({
                    "username":      new_username,
                    "password_hash": bcrypt.hashpw(new_password.encode(), bcrypt.gensalt()).decode(),
                    "display_name":  new_display,
                    "role":          new_role.lower(),
                    "status":        "approved",
                    "requested_at":  datetime.datetime.now().isoformat(),
                    "decided_at":    datetime.datetime.now().isoformat(),
                })
                _save_users(svc, users_file_id, all_users)
                st.success(f"User '{new_username}' added and approved as {new_role}.")
                st.rerun()


title_col, toggle_col, profile_col = st.columns([5, 1, 1])
with title_col:
    st.title("Linchris Hotel Corporation — Weekly Update Tools")
with toggle_col:
    st.write("")
    test_mode = st.toggle("Test Mode", value=False, key="test_mode")
with profile_col:
    st.write("")
    with st.popover(f"👤 {st.session_state.get('username') or 'Account'}"):
        st.write(f"**{st.session_state.get('username')}**")
        st.caption("Admin" if st.session_state.get("is_admin") else "Editor")
        if st.session_state.get("is_admin"):
            if st.button("⚙️ Admin Settings", key="open_admin_settings"):
                st.session_state["view"] = "admin_settings"
                st.rerun()
        if st.button("Log Out", key="logout_btn"):
            st.session_state["authenticated"] = False
            st.session_state["is_admin"] = False
            st.session_state["username"] = None
            st.session_state["view"] = "main"
            st.rerun()

if st.session_state.get("view") == "admin_settings" and st.session_state.get("is_admin"):
    _admin_svc = None
    try:
        _admin_svc = get_drive_service()
        _users_file_id, _users_err = _find_or_create_users_file(_admin_svc)
    except Exception as e:
        _users_file_id, _users_err = None, str(e)
    render_admin_settings(_admin_svc, _users_file_id, _users_err)
    st.stop()

# ── Manual upload (test mode only) ───────────────────────────────────────────
if test_mode:
 with st.expander("Manual Upload", expanded=False):
    with st.expander("ROB Update"):
        st.header("ROB Master Workbook Update")
        csv_file = st.file_uploader("Upload CSV (Business on the Books)", type=["csv", "xlsx"], key="rob_csv")
        xl_file  = st.file_uploader("Upload ROB Master Workbook (.xlsx)", type=["xlsx"], key="rob_xl")
        npu_compare_file = st.file_uploader(
            "Occupancy Statistics — with unpicked group revenue included (Margaritaville only, optional)",
            type=["xlsx"], key="rob_npu_compare")

        if csv_file and xl_file:
            xl_bytes  = xl_file.read()

            df = parse_bob_source(csv_file)
            npu_compare_df = parse_bob_source(npu_compare_file) if npu_compare_file else None
            wb = openpyxl.load_workbook(io.BytesIO(xl_bytes), data_only=False)

            auto_sheet = first_uncolored_sheet(wb, ROB_SHEETS)

            # A selectbox's `index=` is only honored the first time its `key` is
            # created — once rob_sheet exists in session_state, Streamlit ignores
            # index= on every rerun and keeps the old selection. Force a reset
            # whenever the uploaded file's bytes change (see the identical fix
            # applied to the Forecast manual-upload tab).
            rob_xl_hash = hashlib.md5(xl_bytes).hexdigest()
            if st.session_state.get("rob_xl_hash") != rob_xl_hash:
                st.session_state["rob_xl_hash"] = rob_xl_hash
                st.session_state["rob_sheet"] = auto_sheet

            sheet_choice = st.selectbox("Week tab", ROB_SHEETS, key="rob_sheet")
            st.caption(f"Auto-detected next tab: **{auto_sheet}**")
    
            if st.button("Preview Changes", key="rob_preview"):
                ws = wb[sheet_choice]
                grp_npu_rev_override = compute_grp_npu_rev_override(df, npu_compare_df)
                changes = build_rob_change_plan(df, ws, grp_npu_rev_override=grp_npu_rev_override)
                st.session_state["rob_changes"]   = changes
                st.session_state["rob_wb_bytes"]  = xl_bytes
                st.session_state["rob_sheet_sel"] = sheet_choice
    
            if "rob_changes" in st.session_state:
                changes    = st.session_state["rob_changes"]
                will_write = [c for c in changes if not c["skip_reason"]]
                skipped    = [c for c in changes if c["skip_reason"]]
    
                c1, c2 = st.columns(2)
                c1.metric("Cells to update", len(will_write))
                c2.metric("Skipped",         len(skipped))
    
                preview_rows = []
                for c in changes:
                    preview_rows.append({
                        "Month":  c["month"] or "—",
                        "Label":  c["label"],
                        "Row":    c["row"],
                        "Col":    c["col"],
                        "Value":  c["new_value"],
                        "Status": "✅ will write" if not c["skip_reason"] else f"⚠️ skip ({c['skip_reason']})",
                    })
                st.dataframe(preview_rows, use_container_width=True)
    
                if st.button("Confirm and Apply Changes", key="rob_apply"):
                    wb2 = openpyxl.load_workbook(io.BytesIO(st.session_state["rob_wb_bytes"]), data_only=False)
                    apply_rob_changes(wb2, st.session_state["rob_sheet_sel"], changes)
                    color_tab_done(wb2, st.session_state["rob_sheet_sel"])
                    strip_tables(wb2)
                    out = io.BytesIO()
                    wb2.save(out)
                    st.download_button(
                        "Download Updated ROB Workbook",
                        data=out.getvalue(),
                        file_name="ROB_Master_updated.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
    
    with st.expander("Strategy Report"):
        st.header("Strategy Report Update")

        col_a, col_b = st.columns(2)
        with col_a:
            csv_file2 = st.file_uploader("Upload CSV (Business on the Books)", type=["csv", "xlsx"], key="str_csv")
        with col_b:
            rate_file2 = st.file_uploader("Upload Rates & Restrictions CSV", type=["csv"], key="str_rate")
    
        col_c, col_d = st.columns(2)
        with col_c:
            xl_file2 = st.file_uploader("Upload Strategy Report Workbook (.xlsx)", type=["xlsx"], key="str_xl")
    
        if xl_file2:
            xl_bytes2 = xl_file2.read()
            wb2_peek  = openpyxl.load_workbook(io.BytesIO(xl_bytes2), data_only=False)

            auto_sheet2 = first_undone_strategy_sheet(wb2_peek, STRATEGY_SHEETS)

            # Same stale-selection issue as ROB/Forecast: force a reset whenever
            # the uploaded file's bytes change, since index= is ignored once
            # str_sheet already exists in session_state.
            str_xl_hash = hashlib.md5(xl_bytes2).hexdigest()
            if st.session_state.get("str_xl_hash") != str_xl_hash:
                st.session_state["str_xl_hash"] = str_xl_hash
                st.session_state["str_sheet"] = auto_sheet2

            sheet_choice2 = st.selectbox("Week tab", STRATEGY_SHEETS, key="str_sheet")
            st.caption(f"Auto-detected next tab: **{auto_sheet2}**")
    
            if (csv_file2 or rate_file2) and st.button("Preview Changes", key="str_preview"):
                wb2_full = openpyxl.load_workbook(io.BytesIO(xl_bytes2), data_only=False)
                all_changes = []
    
                if csv_file2:
                    df2 = parse_bob_source(csv_file2)
                    all_changes += build_strategy_change_plan(df2, wb2_full, sheet_choice2)
    
                rate_warnings = []
                if rate_file2:
                    rate_df = parse_rate_csv(rate_file2.read())
                    rate_changes, rate_warnings = build_rates_change_plan(
                        rate_df, wb2_full, sheet_choice2)
                    all_changes += rate_changes
    
                st.session_state["str_changes"]   = all_changes
                st.session_state["str_wb_bytes"]  = xl_bytes2
                st.session_state["str_sheet_sel"] = sheet_choice2
                st.session_state["str_warnings"]  = rate_warnings
    
            if "str_changes" in st.session_state:
                for w in st.session_state.get("str_warnings", []):
                    st.warning(w)
    
                changes2    = st.session_state["str_changes"]
                will_write2 = [c for c in changes2 if not c["skip_reason"]]
                skipped2    = [c for c in changes2 if c["skip_reason"]]
    
                c1, c2, c3 = st.columns(3)
                c1.metric("Cells to update", len(will_write2))
                c2.metric("Skipped",         len(skipped2))
                c3.metric("Days in scope",   len({c["date"] for c in changes2}))
    
                preview_rows2 = []
                for c in changes2:
                    preview_rows2.append({
                        "Date":   str(c["date"]),
                        "Label":  c["label"],
                        "Row":    c["row"],
                        "Col":    c["col"],
                        "Value":  c["new_value"],
                        "Status": "✅ will write" if not c["skip_reason"] else f"⚠️ skip ({c['skip_reason']})",
                    })
                st.dataframe(preview_rows2, use_container_width=True)
    
                if st.button("Confirm and Apply Changes", key="str_apply"):
                    wb3 = openpyxl.load_workbook(io.BytesIO(st.session_state["str_wb_bytes"]), data_only=False)
                    apply_strategy_changes(wb3, st.session_state["str_sheet_sel"], changes2)
                    color_tab_done(wb3, st.session_state["str_sheet_sel"])
                    strip_tables(wb3)
                    out2 = io.BytesIO()
                    wb3.save(out2)
                    st.download_button(
                        "Download Updated Strategy Workbook",
                        data=out2.getvalue(),
                        file_name="Strategy_Report_updated.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
    
    with st.expander("Forecast"):
        st.header("Forecast Update")
    
        fcst_csv     = st.file_uploader("Upload CSV (Business on the Books)", type=["csv", "xlsx"], key="fcst_csv")
        fcst_xl      = st.file_uploader("Upload Current Month Forecast Workbook (.xlsx/.xlsm)", type=["xlsx", "xlsm"], key="fcst_xl")
        st.caption("Weeks 3 & 4 only: also upload next month's forecast workbook.")
        fcst_xl_next = st.file_uploader("Upload Next Month Forecast Workbook (.xlsx/.xlsm)", type=["xlsx", "xlsm"], key="fcst_xl_next")
    
        if fcst_xl:
            fcst_bytes = fcst_xl.read()
            wb_fcst_peek = openpyxl.load_workbook(io.BytesIO(fcst_bytes), data_only=False)
    
            avail_fcst = [s for s in FORECAST_SHEETS if s in wb_fcst_peek.sheetnames]
            if not avail_fcst:
                st.error(f"None of the expected week tabs (FCST-WK1 – FCST-WK9) were found in "
                         f"this workbook. It has: {', '.join(wb_fcst_peek.sheetnames)}. Make sure "
                         f"you uploaded the destination Forecast workbook here, not the source "
                         f"data file.")
            auto_fcst  = first_unhighlighted_forecast_sheet(wb_fcst_peek, avail_fcst) if avail_fcst else None

            # A selectbox's `index=` is only honored the first time its `key` is
            # created — once fcst_sheet exists in session_state, Streamlit ignores
            # index= on every rerun and keeps the old selection. Re-uploading a new
            # file (this week's, with more tabs now done) would silently keep
            # writing to whatever tab was last picked instead of the newly
            # auto-detected one. Force a reset whenever the uploaded bytes change.
            fcst_hash = hashlib.md5(fcst_bytes).hexdigest()
            if st.session_state.get("fcst_xl_hash") != fcst_hash:
                st.session_state["fcst_xl_hash"] = fcst_hash
                if auto_fcst:
                    st.session_state["fcst_sheet"] = auto_fcst

            fcst_sheet = st.selectbox("Week tab", avail_fcst, key="fcst_sheet")
            if auto_fcst:
                st.caption(f"Auto-detected next tab: **{auto_fcst}**")
    
            if fcst_xl_next:
                fcst_next_bytes = fcst_xl_next.read()
                wb_next_peek    = openpyxl.load_workbook(io.BytesIO(fcst_next_bytes), data_only=False)
                avail_next      = [s for s in FORECAST_SHEETS if s in wb_next_peek.sheetnames]
                auto_next       = first_unhighlighted_forecast_sheet(wb_next_peek, avail_next) if avail_next else None

                fcst_next_hash = hashlib.md5(fcst_next_bytes).hexdigest()
                if st.session_state.get("fcst_xl_next_hash") != fcst_next_hash:
                    st.session_state["fcst_xl_next_hash"] = fcst_next_hash
                    if auto_next:
                        st.session_state["fcst_sheet_next"] = auto_next

                fcst_sheet_next = st.selectbox("Next month week tab", avail_next, key="fcst_sheet_next")
                if auto_next:
                    st.caption(f"Auto-detected next month tab: **{auto_next}**")
    
            if fcst_csv and fcst_sheet and st.button("Preview Changes", key="fcst_preview"):
                # Current month workbook
                wb_fcst_full = openpyxl.load_workbook(io.BytesIO(fcst_bytes), data_only=False)
                df_fcst = parse_bob_source(fcst_csv)
                ws_fcst = wb_fcst_full[fcst_sheet]
                fcst_changes, fcst_warnings = build_forecast_change_plan(df_fcst, ws_fcst)
    
                st.session_state["fcst_changes"]   = fcst_changes
                st.session_state["fcst_wb_bytes"]  = fcst_bytes
                st.session_state["fcst_sheet_sel"] = fcst_sheet
                st.session_state["fcst_warnings"]  = fcst_warnings
    
                # Next month workbook (weeks 3 & 4)
                if fcst_xl_next:
                    wb_next_full = openpyxl.load_workbook(io.BytesIO(fcst_next_bytes), data_only=False)
                    ws_next      = wb_next_full[fcst_sheet_next]
                    next_changes, next_warnings = build_next_month_forecast_plan(df_fcst, ws_next)
                    st.session_state["fcst_next_changes"]   = next_changes
                    st.session_state["fcst_next_wb_bytes"]  = fcst_next_bytes
                    st.session_state["fcst_next_sheet_sel"] = fcst_sheet_next
                    st.session_state["fcst_next_warnings"]  = next_warnings
                else:
                    st.session_state.pop("fcst_next_changes", None)
    
            if "fcst_changes" in st.session_state:
                for w in st.session_state.get("fcst_warnings", []):
                    st.warning(w)
    
                fcst_ch   = st.session_state["fcst_changes"]
                will_fcst = [c for c in fcst_ch if not c["skip_reason"]]
                skip_fcst = [c for c in fcst_ch if c["skip_reason"]]
    
                st.subheader("Current month changes")
                c1, c2 = st.columns(2)
                c1.metric("Cells to update", len(will_fcst))
                c2.metric("Skipped",         len(skip_fcst))
    
                preview_fcst = []
                for c in fcst_ch:
                    preview_fcst.append({
                        "Label":  c["label"],
                        "Row":    c["row"],
                        "Col":    c["col"],
                        "Value":  c["new_value"],
                        "Status": "✅ will write" if not c["skip_reason"] else f"⚠️ skip ({c['skip_reason']})",
                    })
                st.dataframe(preview_fcst, use_container_width=True)
    
                if "fcst_next_changes" in st.session_state:
                    next_ch = st.session_state["fcst_next_changes"]
                    for w in st.session_state.get("fcst_next_warnings", []):
                        st.warning(w)
                    will_next = [c for c in next_ch if not c["skip_reason"]]
                    skip_next = [c for c in next_ch if c["skip_reason"]]
                    st.subheader("Next month changes")
                    n1, n2 = st.columns(2)
                    n1.metric("Cells to update", len(will_next))
                    n2.metric("Skipped",         len(skip_next))
                    preview_next = []
                    for c in next_ch:
                        preview_next.append({
                            "Label":  c["label"],
                            "Row":    c["row"],
                            "Col":    c["col"],
                            "Value":  c["new_value"],
                            "Status": "✅ will write" if not c["skip_reason"] else f"⚠️ skip ({c['skip_reason']})",
                        })
                    st.dataframe(preview_next, use_container_width=True)
    
                if st.button("Confirm and Apply Changes", key="fcst_apply"):
                    # Apply current month
                    wb_out = openpyxl.load_workbook(io.BytesIO(st.session_state["fcst_wb_bytes"]), data_only=False)
                    apply_forecast_changes(wb_out, st.session_state["fcst_sheet_sel"], fcst_ch)
                    color_tab_done(wb_out, st.session_state["fcst_sheet_sel"])
                    strip_tables(wb_out)
                    out3 = io.BytesIO()
                    wb_out.save(out3)
                    st.download_button(
                        "Download Updated Current Month Forecast",
                        data=out3.getvalue(),
                        file_name="Forecast_current_updated.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
    
                    # Apply next month (if uploaded)
                    if "fcst_next_changes" in st.session_state:
                        wb_next_out = openpyxl.load_workbook(io.BytesIO(st.session_state["fcst_next_wb_bytes"]), data_only=False)
                        apply_forecast_changes(wb_next_out, st.session_state["fcst_next_sheet_sel"],
                                               st.session_state["fcst_next_changes"])
                        color_tab_done(wb_next_out, st.session_state["fcst_next_sheet_sel"])
                        strip_tables(wb_next_out)
                        out4 = io.BytesIO()
                        wb_next_out.save(out4)
                        st.download_button(
                            "Download Updated Next Month Forecast",
                            data=out4.getvalue(),
                            file_name="Forecast_next_month_updated.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
    
st.divider()
# ── Google Drive Update ───────────────────────────────────────────────────────
col_title, col_month = st.columns([5, 2])
with col_title:
    st.header("Weekly Workbook Update")
with col_month:
    st.write("")
    st.write("")
    st.caption(f"📅 Current month: **{datetime.date.today().strftime('%B %Y')}**")

hotels = get_hotels_from_drive()
hotel_names = [h[0] for h in hotels]
hotel_id_map = {h[0]: h[1] for h in hotels}

start_new_month = False
with st.container(border=True):
    col_h, col_ref, col_w = st.columns([3, 1, 3])
    with col_h:
        hotel_sel = st.selectbox("Hotel", hotel_names if hotel_names else ["(no hotels found)"], key="drive_hotel")
    with col_ref:
        st.write("")
        if st.button("↺", key="refresh_hotels", help="Refresh hotel list"):
            get_hotels_from_drive.clear()
            st.rerun()
    with col_w:
        wb_sels = st.pills(
            "Workbooks to update",
            WORKBOOK_TYPES,
            selection_mode="multi",
            default=WORKBOOK_TYPES,
            key="drive_wb",
        ) or []

    # Keying these to the selected hotel clears any uploaded file the moment you
    # switch hotels — one hotel's BOB/R&R CSV should never carry over and get
    # applied to a different hotel.
    drive_csv = st.file_uploader("CSV — Business on the Books", type=["csv", "xlsx"], key=f"drive_csv_{hotel_sel}", width=500)
    drive_rate_csv = None
    if "Strategy Report" in (wb_sels or []):
        drive_rate_csv = st.file_uploader("CSV — Rates & Restrictions", type=["csv"], key=f"drive_rate_csv_{hotel_sel}", width=500)
    drive_npu_compare_csv = None
    if "ROB" in (wb_sels or []) and "margaritaville" in hotel_sel.lower():
        drive_npu_compare_csv = st.file_uploader(
            "Occupancy Statistics — with unpicked group revenue included",
            type=["xlsx"], key=f"drive_npu_compare_{hotel_sel}", width=500)

    opt_col1, opt_col2 = st.columns(2)
    forecast_next_month = False
    if "Forecast" in (wb_sels or []):
        with opt_col1:
            forecast_next_month = st.checkbox("Include next month's Forecast", key="drive_fcst_next")
    with opt_col2:
        start_new_month = st.checkbox("Set up new month", key="drive_new_month")
if start_new_month:
    with st.container(border=True):
        today         = datetime.date.today()
        cur_month_dt  = today.replace(day=1)
        prev_month_dt = (cur_month_dt - datetime.timedelta(days=1)).replace(day=1)
        next_month_dt = (cur_month_dt + datetime.timedelta(days=32)).replace(day=1)
        month_options = {
            prev_month_dt.strftime("%B %Y"): prev_month_dt,
            cur_month_dt.strftime("%B %Y"):  cur_month_dt,
            next_month_dt.strftime("%B %Y"): next_month_dt,
        }
        month_labels = list(month_options.keys())
        sel_month_label = st.selectbox("Month to set up", month_labels,
                                        index=month_labels.index(cur_month_dt.strftime("%B %Y")),
                                        key="setup_month_sel")
        setup_month_dt  = month_options[sel_month_label]
        month_kw        = setup_month_dt.strftime("%b%Y").upper()

        rob_col, sr_col = st.columns(2)

        # ── ROB setup ──────────────────────────────────────────────────────────
        with rob_col:
            st.markdown("**ROB**")
            if st.button("Set Up New ROB", key="btn_setup_rob", type="primary", use_container_width=True):
                try:
                    svc         = get_drive_service()
                    hotel_id_nm = hotel_id_map.get(hotel_sel, "")
                    with st.spinner("Setting up ROB — this may take a moment..."):
                        rob_name, rob_err = setup_new_rob_month(svc, hotel_id_nm, hotel_sel, setup_month_dt)
                    if rob_err and not rob_name:
                        if "storageQuotaExceeded" in str(rob_err):
                            _, master_name = find_rob_master(svc, hotel_id_nm)
                            rob_suffix = hotel_sel.upper()
                            if master_name and "ROB" in master_name.upper():
                                after = master_name[master_name.upper().find("ROB") + 3:].strip()
                                after = after.replace(".xlsx","").replace(".xlsm","").replace(".XLSX","").replace(".XLSM","").strip()
                                if after:
                                    rob_suffix = after
                            ext = ".xlsm" if master_name and master_name.lower().endswith(".xlsm") else ".xlsx"
                            suggested_name = f"{month_kw} ROB {rob_suffix}{ext}"
                            st.warning(
                                f"Auto-copy requires a Shared Drive. Do this in Google Drive first:\n\n"
                                f"1. Right-click **{master_name or 'the ROB master'}** → *Make a copy*\n"
                                f"2. Rename to: **`{suggested_name}`**\n"
                                f"3. Move into the **{month_kw}** folder\n\n"
                                f"Then click **Set Up New ROB** again."
                            )
                        else:
                            st.error(f"ROB setup error: {rob_err}")
                    else:
                        if rob_err:
                            st.warning(rob_err)
                        st.success(f"**{rob_name}** ready for {setup_month_dt.strftime('%B %Y')}.")
                except Exception as e:
                    st.error(f"ROB setup error: {e}")

        # ── Strategy Report setup ──────────────────────────────────────────────
        with sr_col:
            st.markdown("**Strategy Report**")
            if st.button("Set Up New SR", key="btn_setup_new_wb", type="primary", use_container_width=True):
                try:
                    svc         = get_drive_service()
                    hotel_id_nm = hotel_id_map.get(hotel_sel, "")

                    # Step 1 — ensure the file exists; skip copy if it's already there
                    is_fresh_copy = False
                    with st.spinner("Step 1 / 3 — locating or creating workbook..."):
                        existing, find_err = resolve_drive_workbook(svc, hotel_id_nm, hotel_sel,
                                                              "Strategy Report", month_date=setup_month_dt)
                        if existing:
                            st.info(f"Found existing file: **{existing[1]}** — skipping copy.")
                        else:
                            is_fresh_copy = True
                            created_name, create_err = setup_new_sr_month(svc, hotel_id_nm, hotel_sel, setup_month_dt)
                            if create_err:
                                master_id, master_name = find_sr_master(svc, hotel_id_nm)
                                hotel_suffix = ""
                                if master_name and "STRATEGY" in master_name.upper():
                                    hotel_suffix = master_name[master_name.upper().find("STRATEGY") + len("STRATEGY"):].strip().replace(".xlsx","").replace(".XLSX","").strip()
                                suggested_name = f"{month_kw} STRATEGY {hotel_suffix}.xlsx".strip()
                                if "storageQuotaExceeded" in str(create_err):
                                    st.warning(
                                        f"Auto-copy requires a Shared Drive. Do this in Google Drive first:\n\n"
                                        f"1. Right-click **{master_name or 'the SR master'}** → *Make a copy*\n"
                                        f"2. Rename to: **`{suggested_name}`**\n"
                                        f"3. Move into the **{month_kw}** folder\n\n"
                                        f"Then click **Set Up New SR** again."
                                    )
                                else:
                                    st.error(f"Could not create workbook: {create_err}")
                                st.stop()

                    # Step 2 — load reference workbooks into memory
                    with st.spinner("Step 2 / 3 — loading reference workbooks..."):
                        prev_month_dt    = (setup_month_dt - datetime.timedelta(days=1)).replace(day=1)
                        ly_month_dt      = setup_month_dt.replace(year=setup_month_dt.year - 1)
                        prev_month_sr_wb = _load_wb_from_drive(svc, hotel_id_nm, hotel_sel, "Strategy Report", prev_month_dt, data_only=False)
                        ly_sr_wb         = _load_wb_from_drive(svc, hotel_id_nm, hotel_sel, "Strategy Report", ly_month_dt)
                    st.info(f"Prev month ({prev_month_dt.strftime('%b %Y')}): {'✓' if prev_month_sr_wb else '✗ not found'}")
                    st.info(f"Last year  ({ly_month_dt.strftime('%b %Y')}): {'✓' if ly_sr_wb else '✗ not found'}")

                    # Step 3 — populate all 5 weeks
                    with st.spinner("Step 3 / 3 — populating all weeks..."):
                        result, err = resolve_drive_workbook(svc, hotel_id_nm, hotel_sel,
                                                             "Strategy Report", month_date=setup_month_dt)
                        if err:
                            st.error(f"Cannot open new workbook: {err}")
                            st.stop()
                        file_id, file_name = result
                        wb_bytes = drive_download(svc, file_id)
                        original_bytes = wb_bytes
                        wb       = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=False)
                        if is_fresh_copy:
                            clear_tab_colors(wb, STRATEGY_SHEETS)
                        restructure_sr_dates(wb, setup_month_dt)
                        first_ws = wb[STRATEGY_SHEETS[0]] if STRATEGY_SHEETS[0] in wb.sheetnames else None
                        num_rows = _count_sheet_data_rows(first_ws) if first_ws else 365
                        full_scope_start = setup_month_dt
                        full_scope_end   = setup_month_dt + datetime.timedelta(days=max(0, num_rows - 1))
                        total_written = 0
                        for sheet_name in STRATEGY_SHEETS:
                            if sheet_name not in wb.sheetnames:
                                continue
                            changes = build_strategy_change_plan(None, wb, sheet_name,
                                                                  prev_month_wb=prev_month_sr_wb,
                                                                  ly_wb=ly_sr_wb,
                                                                  scope_start=full_scope_start,
                                                                  scope_end=full_scope_end)
                            apply_strategy_changes(wb, sheet_name, changes)
                            total_written += len([c for c in changes if not c.get("skip_reason")])
                        strip_tables(wb)
                        out = io.BytesIO()
                        wb.save(out)
                        drive_upload(svc, file_id, out.getvalue(), file_name)
                        st.session_state["setup_undo"] = {
                            "file_id":   file_id,
                            "file_name": file_name,
                            "bytes":     original_bytes,
                        }

                    st.success(
                        f"**{file_name}** is set up for {setup_month_dt.strftime('%B %Y')}. "
                        f"Populated **{total_written}** cells across all weeks."
                    )
                except Exception as e:
                    st.error(f"Setup error: {e}")

        # Reset button — shown after a successful setup
        if "setup_undo" in st.session_state:
            st.divider()
            reset_col, _ = st.columns([2, 5])
            with reset_col:
             if st.button("↩ Reset Workbook to Original", key="setup_reset", type="secondary", use_container_width=True):
                 try:
                     info = st.session_state["setup_undo"]
                     with st.spinner("Restoring original workbook..."):
                         get_drive_service()
                         drive_upload(get_drive_service(), info["file_id"], info["bytes"], info["file_name"])
                     del st.session_state["setup_undo"]
                     st.success(f"**{info['file_name']}** restored to original state.")
                 except Exception as e:
                     st.error(f"Reset error: {e}")




def build_all_plans(svc, hotel_sel, hotel_id, wb_sels, df, rate_df, forecast_next_month=False, npu_compare_df=None):
    today = datetime.date.today()
    current_month = today.replace(day=1)
    all_plans = {}

    grp_npu_rev_override = compute_grp_npu_rev_override(df, npu_compare_df)

    # Pre-load reference workbooks into memory once — used for cross-sheet lookups
    prev_month_sr_wb = None
    ly_sr_wb         = None
    if "Strategy Report" in wb_sels:
        prev_month_dt = (current_month - datetime.timedelta(days=1)).replace(day=1)
        ly_month_dt   = current_month.replace(year=current_month.year - 1)
        prev_month_sr_wb = _load_wb_from_drive(svc, hotel_id, hotel_sel, "Strategy Report", prev_month_dt, data_only=False)
        ly_sr_wb         = _load_wb_from_drive(svc, hotel_id, hotel_sel, "Strategy Report", ly_month_dt)
        # Comp Set LY / OTB LY Trans / GRP LY etc. all come from ly_sr_wb — if it's
        # not found, those fields silently produce nothing (no warning previously),
        # which looked like "dates transferred but no text" with no explanation why.
        st.info(f"SR reference workbooks — Prev month ({prev_month_dt.strftime('%b %Y')}): "
                f"{'✓ found' if prev_month_sr_wb else '✗ NOT FOUND — OTB Lst Wek will be blank'} | "
                f"Last year ({ly_month_dt.strftime('%b %Y')}): "
                f"{'✓ found' if ly_sr_wb else '✗ NOT FOUND — all LY columns (incl. Comp Set LY text) will be blank'}")

    for wb_type in wb_sels:
        result, err = resolve_drive_workbook(svc, hotel_id, hotel_sel, wb_type)
        if err:
            st.error(f"{wb_type}: {err}")
            continue
        file_id, file_name = result
        wb_bytes = drive_download(svc, file_id)
        wb       = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=False)
        if wb_type == "ROB":
            avail    = [s for s in ROB_SHEETS if s in wb.sheetnames]
            auto     = first_uncolored_sheet(wb, avail)
            sheet    = auto or avail[0]
            changes  = build_rob_change_plan(df, wb[sheet], grp_npu_rev_override=grp_npu_rev_override)
            warnings = []
        elif wb_type == "Strategy Report":
            avail    = [s for s in STRATEGY_SHEETS if s in wb.sheetnames]
            auto     = first_undone_strategy_sheet(wb, avail)
            sheet    = auto or avail[0]
            date_row_map_debug = build_date_row_map(wb)
            st.info(f"SR: **{file_name}** → sheet **{sheet}** | "
                    f"date rows mapped: {len(date_row_map_debug)} | "
                    f"date range: {min(date_row_map_debug) if date_row_map_debug else 'none'} – {max(date_row_map_debug) if date_row_map_debug else 'none'}")
            if df is not None:
                sample_dates = [str(df.iloc[i, 0]) for i in range(min(5, len(df)))]
                bob_daily = sum(1 for _, r in df.iterrows() if classify_row(str(r[0]).strip())[0] == "daily")
                st.info(f"BOB CSV: {len(df)} rows | daily rows matched: {bob_daily} | first 5 col-0 values: {sample_dates}")
            else:
                st.warning("BOB CSV: df is None — no CSV uploaded or parse failed")
            changes  = build_strategy_change_plan(df, wb, sheet,
                                                   prev_month_wb=prev_month_sr_wb,
                                                   ly_wb=ly_sr_wb)
            warnings = []
            if rate_df is not None:
                rate_changes, rate_warnings = build_rates_change_plan(rate_df, wb, sheet)
                changes  += rate_changes
                warnings += rate_warnings
        else:  # Forecast — current month (no Month Ending Forecast fill here)
            avail    = [s for s in FORECAST_SHEETS if s in wb.sheetnames]
            auto     = first_unhighlighted_forecast_sheet(wb, avail)
            sheet    = auto or avail[0]
            changes, warnings = build_forecast_change_plan(df, wb[sheet])
        all_plans[wb_type] = {
            "file_id":   file_id,
            "file_name": file_name,
            "wb_bytes":  wb_bytes,
            "sheet":     sheet,
            "changes":   changes,
            "warnings":  warnings,
        }

    # Next-month Forecast: only when checkbox is ticked
    if "Forecast" in wb_sels and forecast_next_month:
        next_month_dt = (datetime.date.today().replace(day=1) + datetime.timedelta(days=32)).replace(day=1)
        nm_result, nm_err = resolve_drive_workbook(svc, hotel_id, hotel_sel, "Forecast", month_date=next_month_dt)
        if nm_err:
            # Workbook not found — auto-create from master
            st.info(f"Next month Forecast not found — creating from master...")
            created_name, setup_err = setup_new_forecast_month(svc, hotel_id, hotel_sel, next_month_dt)
            if setup_err and not created_name:
                st.warning(f"Next month Forecast: {setup_err}")
            else:
                if setup_err:
                    st.warning(setup_err)
                else:
                    st.success(f"Created **{created_name}** for {next_month_dt.strftime('%B %Y')}.")
                nm_result, nm_err = resolve_drive_workbook(svc, hotel_id, hotel_sel, "Forecast", month_date=next_month_dt)
                if nm_err:
                    st.warning(f"Still could not find next month Forecast after creation: {nm_err}")
        else:
            nm_file_id, nm_file_name = nm_result
            nm_bytes = drive_download(svc, nm_file_id)
            nm_wb    = openpyxl.load_workbook(io.BytesIO(nm_bytes), data_only=False)
            nm_avail = [s for s in FORECAST_SHEETS if s in nm_wb.sheetnames]
            nm_auto  = first_unhighlighted_forecast_sheet(nm_wb, nm_avail)
            nm_sheet = nm_auto or nm_avail[0]
            nm_changes, nm_warnings = build_next_month_forecast_plan(df, nm_wb[nm_sheet])
            # Month Ending Forecast table — fill Budget + LY from next month's ROB
            nm_is_wk1 = (nm_sheet == FORECAST_SHEETS[0])
            if nm_is_wk1:
                nm_rob_result, _ = resolve_drive_workbook(svc, hotel_id, hotel_sel, "ROB",
                                                           month_date=next_month_dt)
                if nm_rob_result:
                    nm_rob_wb = openpyxl.load_workbook(
                        io.BytesIO(drive_download(svc, nm_rob_result[0])), data_only=True)
                    extra, extra_warn = build_forecast_change_plan(
                        df, nm_wb[nm_sheet], rob_wb=nm_rob_wb, is_wk1=True)
                    # Only keep the Month Ending Forecast entries from extra
                    nm_changes += [c for c in extra if "Month End Forecast" in c.get("label", "")]
                    nm_warnings += extra_warn
            all_plans["Forecast (next month)"] = {
                "file_id":   nm_file_id,
                "file_name": nm_file_name,
                "wb_bytes":  nm_bytes,
                "sheet":     nm_sheet,
                "changes":   nm_changes,
                "warnings":  nm_warnings,
            }

    return all_plans


def _snapshot_changes(wb, sheet_name, changes):
    """Return {(sheet, row, col): original_value} for every cell in the change plan."""
    ws = wb[sheet_name]
    return {
        (sheet_name, ch["row"], ch["col"]): ws.cell(ch["row"], ch["col"]).value
        for ch in changes
        if not ch.get("skip_reason")
    }


def apply_and_upload(svc, all_plans):
    saved, errors = [], []
    undo_snapshot = {}  # cumulative snapshot across all workbooks
    for wb_type, plan in all_plans.items():
        try:
            wb_apply = openpyxl.load_workbook(io.BytesIO(plan["wb_bytes"]), data_only=False)
            # Snapshot originals BEFORE writing
            snap = _snapshot_changes(wb_apply, plan["sheet"], plan["changes"])
            undo_snapshot[wb_type] = {
                "file_id":   plan["file_id"],
                "file_name": plan["file_name"],
                "wb_bytes":  plan["wb_bytes"],   # clean pre-write bytes
                "sheet":     plan["sheet"],
                "cells":     snap,
            }
            if wb_type == "ROB":
                apply_rob_changes(wb_apply, plan["sheet"], plan["changes"])
            elif wb_type == "Strategy Report":
                apply_strategy_changes(wb_apply, plan["sheet"], plan["changes"])
            else:
                apply_forecast_changes(wb_apply, plan["sheet"], plan["changes"])
            color_tab_done(wb_apply, plan["sheet"])
            strip_tables(wb_apply)
            out = io.BytesIO()
            wb_apply.save(out)
            drive_upload(svc, plan["file_id"], out.getvalue(), plan["file_name"])
            saved.append(plan["file_name"])
        except Exception as e:
            errors.append(f"{wb_type}: {e}")
    if saved:
        st.session_state["undo_snapshot"] = undo_snapshot
    return saved, errors


def undo_all_changes(svc):
    """Restore every snapshotted cell to its original value and re-upload."""
    snapshot = st.session_state.get("undo_snapshot", {})
    if not snapshot:
        return [], ["No snapshot found — nothing to undo."]
    saved, errors = [], []
    for wb_type, info in snapshot.items():
        try:
            wb = openpyxl.load_workbook(io.BytesIO(info["wb_bytes"]), data_only=False)
            ws = wb[info["sheet"]]
            for (sheet, row, col), orig_val in info["cells"].items():
                ws.cell(row, col).value = orig_val
            strip_tables(wb)
            out = io.BytesIO()
            wb.save(out)
            drive_upload(svc, info["file_id"], out.getvalue(), info["file_name"])
            saved.append(info["file_name"])
        except Exception as e:
            errors.append(f"{wb_type}: {e}")
    if saved:
        del st.session_state["undo_snapshot"]
    return saved, errors


ready = drive_csv and wb_sels

if test_mode:
    # ── Test mode: preview first, then confirm ────────────────────────────────
    if ready and st.button("Preview Changes", key="drive_preview"):
        try:
            svc     = get_drive_service()
            df      = parse_bob_source(drive_csv) if drive_csv else None
            rate_df = parse_rate_csv(drive_rate_csv.read()) if drive_rate_csv else None
            npu_compare_df = parse_bob_source(drive_npu_compare_csv) if drive_npu_compare_csv else None
            st.session_state["drive_plans"]     = build_all_plans(svc, hotel_sel, hotel_id_map.get(hotel_sel, ""), wb_sels, df, rate_df, forecast_next_month, npu_compare_df)
            st.session_state["drive_hotel_sel"] = hotel_sel
        except Exception as e:
            st.error(f"Drive error: {e}")

    if "drive_plans" in st.session_state:
        all_plans = st.session_state["drive_plans"]
        for wb_type, plan in all_plans.items():
            st.subheader(wb_type)
            st.caption(f"File: **{plan['file_name']}** — Tab: **{plan['sheet']}**")
            for w in plan["warnings"]:
                st.warning(w)
            ch = plan["changes"]
            will_write = [c for c in ch if not c["skip_reason"]]
            skipped    = [c for c in ch if c["skip_reason"]]
            c1, c2 = st.columns(2)
            c1.metric("Cells to update", len(will_write))
            c2.metric("Skipped",         len(skipped))
            st.dataframe([{
                "Label":  c["label"],
                "Row":    c["row"],
                "Col":    c["col"],
                "Value":  c["new_value"],
                "Status": "✅ will write" if not c["skip_reason"] else f"⚠️ skip ({c['skip_reason']})",
            } for c in ch], use_container_width=True)

        if st.button("Confirm and Save All to Google Drive", key="drive_apply"):
            try:
                saved, errors = apply_and_upload(get_drive_service(), all_plans)
                for name in saved:
                    st.success(f"Saved **{name}** to Google Drive.")
                for err in errors:
                    st.error(err)
            except Exception as e:
                st.error(f"Drive error: {e}")
else:
    # ── Normal mode: one click ────────────────────────────────────────────────
    if ready and st.button("Upload Data to Workbooks", key="drive_go", type="primary"):
        try:
            svc     = get_drive_service()
            df      = parse_bob_source(drive_csv) if drive_csv else None
            rate_df = parse_rate_csv(drive_rate_csv.read()) if drive_rate_csv else None
            npu_compare_df = parse_bob_source(drive_npu_compare_csv) if drive_npu_compare_csv else None
            with st.spinner("Updating workbooks in Google Drive..."):
                all_plans       = build_all_plans(svc, hotel_sel, hotel_id_map.get(hotel_sel, ""), wb_sels, df, rate_df, forecast_next_month, npu_compare_df)
                saved, errors   = apply_and_upload(svc, all_plans)
            for name in saved:
                st.success(f"Saved **{name}** to Google Drive.")
            for err in errors:
                st.error(err)
        except Exception as e:
            st.error(f"Drive error: {e}")

# ── Undo button (shown whenever a snapshot exists) ────────────────────────────
if "undo_snapshot" in st.session_state:
    st.divider()
    undo_col, _ = st.columns([2, 5])
    with undo_col:
        if st.button("↩ Undo Last Upload", key="undo_all", type="secondary", use_container_width=True):
            try:
                with st.spinner("Restoring original values..."):
                    saved, errors = undo_all_changes(get_drive_service())
                for name in saved:
                    st.success(f"Restored **{name}** to original state.")
                for err in errors:
                    st.error(err)
            except Exception as e:
                st.error(f"Undo error: {e}")

